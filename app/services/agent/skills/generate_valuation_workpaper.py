"""Skill: chain produce_valuation_inputs → export_workpaper to produce the
populated xlsx. Returns a downloadable URL.
"""

from __future__ import annotations

import asyncio
import json
import logging
import re
import sys
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)


def _slugify(name: str) -> str:
    """Make a filename-safe slug from a company name. Lowercase, kebab-case,
    ASCII-only. Empty input falls back to 'company'."""
    if not name:
        return "company"
    s = name.lower()
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return s or "company"


def _overlay_xlsx_authoritative_values(summary: dict[str, Any], xlsx_path: Path) -> None:
    """Read cached cell values from the xlsx's `Valuation Summary` sheet (which
    LibreOffice recalc'd at the end of export()) and overlay them onto the
    Python-computed summary so the dashboard matches the analyst-visible
    workpaper. Silent no-op if cells aren't cached (e.g., recalc was skipped)."""
    try:
        from openpyxl import load_workbook  # type: ignore
        wb = load_workbook(xlsx_path, data_only=True, read_only=True)
        if "Valuation Summary" not in wb.sheetnames:
            return
        vs = wb["Valuation Summary"]

        def fnum(r: int, c: int) -> float | None:
            v = vs.cell(row=r, column=c).value
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                return float(v)
            return None

        # Row layout from build_valuation_summary_formulas:
        #   r7=EV, r8=Equity pre, r9=after DLOM, r10=after DLOC, r11=client interest
        #   r16=per-share basic, r18=per-share diluted, r21=WACC
        # Col C=Per-Management, Col D=Independent
        pm_ev, in_ev = fnum(7, 3), fnum(7, 4)
        pm_eq_pre, in_eq_pre = fnum(8, 3), fnum(8, 4)
        pm_eq_dlom, in_eq_dlom = fnum(9, 3), fnum(9, 4)
        pm_eq_dloc, in_eq_dloc = fnum(10, 3), fnum(10, 4)
        pm_ps_b, in_ps_b = fnum(16, 3), fnum(16, 4)
        pm_ps_d, in_ps_d = fnum(18, 3), fnum(18, 4)
        pm_wacc, in_wacc = fnum(21, 3), fnum(21, 4)

        dcf = summary.setdefault("dcf", {})
        dcf_pm = dcf.setdefault("per_management", {})
        dcf_in = dcf.setdefault("independent", {})
        if pm_ev is not None: dcf_pm["ev"] = pm_ev
        if in_ev is not None: dcf_in["ev"] = in_ev

        bridge = summary.setdefault("bridge", {})
        br_pm = bridge.setdefault("per_management", {})
        br_in = bridge.setdefault("independent", {})
        if pm_eq_pre is not None: br_pm["pre_discount"] = pm_eq_pre
        if in_eq_pre is not None: br_in["pre_discount"] = in_eq_pre
        if pm_eq_dlom is not None: br_pm["after_dlom"] = pm_eq_dlom
        if in_eq_dlom is not None: br_in["after_dlom"] = in_eq_dlom
        if pm_eq_dloc is not None: br_pm["after_dloc"] = pm_eq_dloc
        if in_eq_dloc is not None: br_in["after_dloc"] = in_eq_dloc

        per_share = summary.setdefault("per_share", {})
        ps_pm = per_share.setdefault("per_management", {})
        ps_in = per_share.setdefault("independent", {})
        if pm_ps_b is not None: ps_pm["basic"] = pm_ps_b
        if in_ps_b is not None: ps_in["basic"] = in_ps_b
        if pm_ps_d is not None: ps_pm["diluted"] = pm_ps_d
        if in_ps_d is not None: ps_in["diluted"] = in_ps_d

        wacc = summary.setdefault("wacc", {})
        w_pm = wacc.setdefault("per_management", {})
        w_in = wacc.setdefault("independent", {})
        if pm_wacc is not None: w_pm["wacc"] = pm_wacc
        if in_wacc is not None: w_in["wacc"] = in_wacc

        # Keep the DCF-primary conclusion + cross-check variances consistent
        # with the authoritative xlsx EV (compute.py derived them from the
        # Python EV, which can drift ~1-2% from the recalc'd workbook).
        if pm_ev is not None:
            concluded = summary.get("concluded")
            if isinstance(concluded, dict):
                concluded["ev"] = pm_ev
            cross = summary.get("cross_checks")
            if isinstance(cross, dict):
                cross["primary_ev"] = pm_ev
                tol = float(cross.get("tolerance_pct") or 0.10)
                any_available = False
                any_outside = False
                for chk in cross.get("checks") or []:
                    if not chk.get("available"):
                        continue
                    implied = chk.get("implied_ev")
                    if not implied or not pm_ev:
                        continue
                    variance = (float(implied) - pm_ev) / pm_ev
                    chk["variance_pct"] = variance
                    chk["within_range"] = abs(variance) <= tol
                    any_available = True
                    any_outside = any_outside or not chk["within_range"]
                if any_available:
                    cross["verdict"] = ("outside_cross_check_range" if any_outside
                                        else "within_reasonable_range")
    except Exception:
        # Non-fatal — Python-computed summary stays in place.
        pass

from app.config import settings
from app.services.agent.context import AgentContext
from app.services.agent.registry import registry
from app.services.agent.skill import Skill, SkillResult, SkillStatus


# Resolve from backend/ root (parents[4]) so paths hold in deploys that ship
# only the backend tree. materials/templates/ and the valuation/ module both
# live inside backend/ now — see also produce_valuation_inputs.py.
BACKEND_ROOT = Path(__file__).resolve().parents[4]
DEFAULT_SKELETON = BACKEND_ROOT / "materials" / "templates" / "orionmano-valuation-template-v1.xlsx"

# Make the standalone valuation module importable
_VAL_DIR = BACKEND_ROOT / "valuation"
if str(_VAL_DIR) not in sys.path:
    sys.path.insert(0, str(_VAL_DIR))


class GenerateValuationWorkpaperSkill(Skill):
    name = "generate_valuation_workpaper"
    description = (
        "Generate a populated valuation workpaper (xlsx) for the company. "
        "Chains produce_valuation_inputs → export_workpaper. Returns a downloadable URL."
    )
    parameters = []

    async def execute(self, ctx: AgentContext, **kwargs: Any) -> SkillResult:
        # 1a. If the caller passed an explicit `inputs` payload (user-edited JSON
        # re-uploaded after a prior run — Eric 2026-05-08 item 8), skip the LLM
        # producer entirely and use the uploaded inputs verbatim. The endpoint
        # should have validated against the schema before passing them in.
        inputs_override = kwargs.get("inputs")
        producer_token_usage = 0
        if inputs_override is not None:
            if not isinstance(inputs_override, dict):
                return SkillResult.failed("Override inputs must be a JSON object")
            payload = inputs_override
            # The producer normally loads ctx.company; this path skips it,
            # which left the output filename slug falling back to "company"
            # and the linked report tier falling back to "standard".
            try:
                await ctx.load_company_data()
            except Exception:
                pass
        else:
            # 1b. Standard path: produce inputs JSON via the producer skill.
            producer = registry.get("produce_valuation_inputs")
            if producer is None:
                return SkillResult.failed("produce_valuation_inputs skill not registered")

            # Forward per-valuation run-config kwargs (e.g. target_valuation from
            # the UI run-config box) so the producer can inject them into the LLM
            # context and authoritatively override the payload after validation.
            producer_result = await producer.execute(ctx, **kwargs)
            if producer_result.status != SkillStatus.SUCCESS:
                return SkillResult.failed(
                    f"Producer skill failed: {producer_result.message}",
                    data=producer_result.data,
                )
            payload = producer_result.data
            if not isinstance(payload, dict):
                return SkillResult.failed("Producer skill returned non-dict payload")
            producer_token_usage = producer_result.token_usage

        # 2. Resolve output path under the configured upload dir, exposed at /uploads/...
        upload_root = Path(settings.UPLOAD_DIR).resolve()
        out_dir = upload_root / "valuations"
        out_dir.mkdir(parents=True, exist_ok=True)

        company_name = getattr(ctx.company, "name", None) if ctx.company else None
        slug = _slugify(company_name or "")
        date_str = datetime.utcnow().strftime("%d%m%Y")
        # If multiple workpapers in one day, suffix with -2, -3, etc. so we
        # never overwrite earlier runs.
        base = f"valuation-{slug}-{date_str}"
        output_path = out_dir / f"{base}.xlsx"
        n = 2
        while output_path.exists():
            output_path = out_dir / f"{base}-{n}.xlsx"
            n += 1

        # 3. Run the export pipeline
        # Skeleton may not exist yet — build_skeleton.py auto-builds when missing
        if not DEFAULT_SKELETON.exists():
            try:
                from build_skeleton import build as build_skeleton  # type: ignore
                build_skeleton()
            except Exception as e:
                return SkillResult.failed(f"Failed to build skeleton: {e}")

        # Normalise the IN-MEMORY payload before serialising. export() runs the
        # same repair, but on the copy it loads from the temp file — so the
        # workbook got scaled Y0 values while the summary, the dashboard and
        # the pinned-parameter prefill all kept the raw ones. That divergence
        # is how an analyst came to pin gross_profit_y0 = 858000 (raw dollars)
        # when the workpaper unit was '000. Normalising here keeps every
        # downstream consumer on the same numbers; export()'s own call is then
        # idempotent.
        try:
            from export_workpaper import normalize_payload, ValidationResult  # type: ignore
            _norm_vr = ValidationResult()
            normalize_payload(payload, _norm_vr)
            for _w in _norm_vr.warnings:
                logger.warning("valuation payload normalised: %s", _w)
        except Exception as e:  # normalisation is a repair, not a gate
            logger.warning("payload normalisation skipped: %s: %s", type(e).__name__, e)

        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".json", delete=False
        ) as f:
            json.dump(payload, f)
            json_path = Path(f.name)

        try:
            from export_workpaper import export  # type: ignore
            vr = export(json_path, DEFAULT_SKELETON, output_path)
        except FileNotFoundError as e:
            return SkillResult.failed(f"Export pipeline missing file: {e}")
        except Exception as e:
            return SkillResult.failed(f"Export failed: {type(e).__name__}: {e}")
        finally:
            json_path.unlink(missing_ok=True)

        # Compute summary + persist alongside the xlsx so the dashboard endpoint
        # can fetch the latest run without rerunning Claude.
        summary: dict[str, Any] | None = None
        summary_path: Path | None = None
        summary_payload: dict[str, Any] | None = None
        try:
            from compute import compute_summary  # type: ignore
            summary = compute_summary(payload)
            # Overlay xlsx-authoritative cached values (LibreOffice recalc'd
            # them inside export()) so the dashboard agrees with the workpaper.
            # If recalc was skipped (LibreOffice unavailable), the overlay
            # leaves summary as Python-computed — analyst will see ~1-2% drift.
            _overlay_xlsx_authoritative_values(summary, output_path)
            summary_payload = {
                "company_id": str(ctx.company_id) if ctx.company_id else None,
                "generated_at": datetime.utcnow().isoformat() + "Z",
                "xlsx_url": f"/uploads/valuations/{output_path.name}",
                "xlsx_filename": output_path.name,
                "warnings": vr.warnings,
                "errors": vr.errors,
                "summary": summary,
                "inputs": payload,
                # Eric 2026-05-19 — only "active" runs surface via /latest and
                # are downloadable. The API endpoint flips the previous active
                # run to "superseded" at the start of /generate-workpaper, so
                # mid-regenerate the user can't accidentally download stale
                # data thinking it's the new run.
                "status": "active",
            }
            summary_path = output_path.with_suffix(".summary.json")
            summary_path.write_text(json.dumps(summary_payload, default=str))
        except Exception as e:
            # Summary failure shouldn't block the workpaper download
            summary = {"error": f"Summary computation failed: {type(e).__name__}: {e}"}

        # Eric 2026-05-13 — every workpaper deliverable must be accompanied by a
        # written report explaining its assumptions and inputs. Create the
        # Report row + schedule background generation; the report reads the
        # .summary.json we just wrote for its authoritative context. Wrapped in
        # try/except so a Report-creation failure doesn't fail the workpaper.
        report_id: str | None = None
        if (
            getattr(ctx, "db", None) is not None
            and getattr(ctx, "company_id", None) is not None
            and getattr(ctx, "user_id", None) is not None
            and summary is not None
            and not (isinstance(summary, dict) and "error" in summary)
        ):
            try:
                from app.models.report import Report
                tier = getattr(ctx.company, "report_tier", None) or "standard"
                company_name = getattr(ctx.company, "name", None) or "Valuation"
                report = Report(
                    company_id=ctx.company_id,
                    report_type="valuation_report",
                    tier=tier,
                    title=f"{company_name} — Valuation Report",
                    status="pending",
                    created_by=ctx.user_id,
                )
                ctx.db.add(report)
                await ctx.db.commit()
                await ctx.db.refresh(report)
                report_id = str(report.id)

                # Surface the linked report in the .summary.json so the dashboard
                # can render a 'View Report' link bound to this workpaper run.
                if summary_payload is not None and summary_path is not None:
                    summary_payload["report_id"] = report_id
                    summary_path.write_text(json.dumps(summary_payload, default=str))

                # Schedule generation on a fresh DB session so the bg task isn't
                # tied to the request-scoped session lifetime.
                rid_uuid = report.id
                cid_uuid = ctx.company_id
                async def _kickoff_report() -> None:
                    from app.database import async_session
                    from app.services.report.generator import generate_report_bg
                    try:
                        async with async_session() as session:
                            await generate_report_bg(
                                session, cid_uuid, "valuation_report", rid_uuid,
                            )
                    except Exception:
                        # Bg task failures shouldn't crash the worker; the Report
                        # row's status will reflect the failure if the generator
                        # got far enough to set it.
                        pass
                asyncio.create_task(_kickoff_report())
            except Exception:
                # Reset the session state if the Report-row write failed mid-flight
                try:
                    await ctx.db.rollback()
                except Exception:
                    pass
                report_id = None

        if vr.errors:
            return SkillResult(
                status=SkillStatus.PARTIAL,
                data={
                    "xlsx_path": str(output_path),
                    "xlsx_url": f"/uploads/valuations/{output_path.name}",
                    "errors": vr.errors,
                    "warnings": vr.warnings,
                    "inputs_json": payload,
                    "summary": summary,
                    "report_id": report_id,
                },
                message=(
                    f"Workpaper generated with {len(vr.errors)} validation errors "
                    f"and {len(vr.warnings)} warnings"
                ),
                artifacts={"xlsx_path": str(output_path), "valuation_inputs": payload},
                token_usage=producer_token_usage,
            )

        return SkillResult.success(
            data={
                "xlsx_path": str(output_path),
                "xlsx_url": f"/uploads/valuations/{output_path.name}",
                "warnings": vr.warnings,
                "inputs_json": payload,
                "summary": summary,
                "report_id": report_id,
            },
            message=(
                f"Workpaper generated at {output_path.name} "
                f"({len(vr.warnings)} warnings)"
                + (f"; written report {report_id} kicked off" if report_id else "")
            ),
            artifacts={"xlsx_path": str(output_path), "valuation_inputs": payload},
        )
