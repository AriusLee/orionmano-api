"""Build the v1 Orionmano valuation workpaper skeleton xlsx.

Deterministic: re-running produces an identical workbook. The Inputs sheet is
populated with the parameter band per the schema in
knowledge-base/04-valuation/inputs-sheet-schema.md, every scalar parameter gets
a workbook-level named range, and computational sheets (Projections, WACC,
DCF, Adjustments, Valuation Summary) ship with formulas pre-wired — they read
their inputs from named ranges so the workbook recalculates when any Inputs
value changes.

Run:
    python3 backend/valuation/build_skeleton.py

Output:
    materials/templates/orionmano-valuation-template-v1.xlsx
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

REPO_ROOT = Path(__file__).resolve().parents[2]
OUTPUT = REPO_ROOT / "materials" / "templates" / "orionmano-valuation-template-v1.xlsx"


# ---------------------------------------------------------------------------
# Schema
# ---------------------------------------------------------------------------

@dataclass
class Param:
    pid: str
    label: str
    type_: str  # text | number | percentage | currency | date | enum | boolean
    scenario: bool = False  # if True, has Per-Mgmt + Independent columns
    notes: str = ""


@dataclass
class Section:
    code: str  # A, B, C, ...
    title: str
    tier: str  # Core | Extended
    params: list[Param] = field(default_factory=list)


# Year-vector helper
def vec(stem: str, label_stem: str, type_: str, n: int = 5) -> list[Param]:
    return [Param(f"{stem}_y{i}", f"{label_stem} (Y{i})", type_) for i in range(1, n + 1)]


SECTIONS: list[Section] = [
    Section("A", "Engagement metadata", "Core", [
        Param("company_name", "Company name", "text"),
        Param("company_country", "Country of incorporation", "enum"),
        Param("company_industry_us", "Industry (US classification)", "enum"),
        Param("company_industry_global", "Industry (Global classification)", "enum"),
        Param("valuation_date", "Valuation date", "date"),
        Param("report_purpose", "Report purpose", "enum"),
        Param("accounting_standard", "Accounting standard", "enum"),
        Param("engagement_team_partner", "Engagement partner", "text"),
        Param("engagement_team_manager", "Engagement manager", "text"),
        Param("engagement_department", "Department", "text"),
        Param("client_name", "Client (engaging party)", "text"),
    ]),
    Section("B", "Currency & units", "Core", [
        Param("currency", "Reporting currency", "enum"),
        Param("unit", "Unit of presentation", "enum"),
        Param("currency_alt", "Secondary display currency (optional)", "enum"),
        Param("fx_rate_alt", "FX rate to alternate currency", "number"),
    ]),
    Section("C", "Tax rule", "Core", [
        Param("tax_jurisdiction", "Tax jurisdiction", "enum"),
        Param("tax_type", "Tax structure", "enum", notes="flat | two_tier | progressive"),
        Param("tax_rate_low", "Low-tier rate", "percentage"),
        Param("tax_rate_high", "High-tier rate (or flat rate)", "percentage"),
        Param("tax_threshold", "Threshold between tiers", "currency"),
        Param("tax_effective_rate", "Effective rate (override)", "percentage"),
    ]),
    Section("D", "Projection drivers", "Core", [
        Param("projection_years", "Explicit forecast years", "number"),
        Param("revenue_growth_method", "Revenue growth method", "enum",
              notes="flat | declining | staged | per_year"),
        *vec("revenue_growth", "Revenue growth", "percentage"),
        *vec("gross_margin", "Gross margin", "percentage"),
        *vec("opex_pct_revenue", "Opex % of revenue", "percentage"),
        *vec("capex_pct_revenue", "Capex % of revenue", "percentage"),
        *vec("dep_pct_revenue", "Depreciation % of revenue", "percentage"),
        *vec("nwc_pct_sales", "Working capital % of sales", "percentage"),
    ]),
    Section("E", "Terminal value", "Core", [
        Param("terminal_method", "Terminal value method", "enum",
              notes="gordon_growth | exit_multiple"),
        Param("terminal_growth_rate", "Terminal growth rate", "percentage"),
        Param("terminal_exit_multiple_type", "Exit multiple metric", "enum"),
        Param("terminal_exit_multiple_value", "Exit multiple value", "number"),
    ]),
    Section("F", "WACC inputs", "Core", [
        Param("risk_free_rate", "Risk-free rate", "percentage"),
        Param("risk_free_rate_source", "Risk-free rate source descriptor", "text"),
        Param("equity_risk_premium", "Equity risk premium (ERP)", "percentage"),
        Param("country_risk_premium", "Country risk premium", "percentage"),
        Param("unlevered_beta", "Unlevered beta", "number", scenario=True),
        Param("target_debt_to_equity", "Target D/E ratio", "percentage", scenario=True),
        Param("levered_beta", "Levered beta (calculated)", "number", scenario=True),
        Param("size_premium", "Size premium", "percentage", scenario=True),
        Param("specific_risk_premium", "Specific risk premium", "percentage", scenario=True),
        Param("cost_of_equity", "Cost of equity (calculated)", "percentage", scenario=True),
        Param("pretax_cost_of_debt", "Pre-tax cost of debt", "percentage", scenario=True),
        Param("aftertax_cost_of_debt", "After-tax cost of debt (calculated)", "percentage", scenario=True),
        Param("target_debt_weight", "D/V (debt weight)", "percentage", scenario=True),
        Param("target_equity_weight", "E/V (equity weight)", "percentage", scenario=True),
        Param("wacc", "WACC (calculated)", "percentage", scenario=True),
    ]),
    Section("I", "EV -> Equity bridge", "Core", [
        Param("surplus_assets", "Surplus / non-operating assets", "currency"),
        Param("net_debt_override", "Net debt (override)", "currency"),
        Param("minority_interests", "Minority interests", "currency"),
        Param("non_operating_assets", "Non-operating assets", "currency"),
        Param("dlom_pct", "DLOM rate", "percentage"),
        Param("dloc_pct", "DLOC rate", "percentage"),
        Param("equity_interest_pct", "Client's % equity interest", "percentage"),
        Param("shares_outstanding", "Shares outstanding (basic)", "number"),
        Param("shares_outstanding_diluted", "Shares outstanding (diluted)", "number"),
        Param("pre_money_pct", "Pre-money equity %", "percentage"),
    ]),
    Section("J", "Adjustment toggles", "Extended", [
        Param("capitalize_rd", "Capitalize R&D?", "boolean"),
        Param("rd_amortization_years", "R&D amortization period (years)", "number"),
        Param("convert_operating_leases", "Convert operating leases to debt?", "boolean"),
        Param("lease_discount_rate", "Lease discount rate", "percentage"),
    ]),
    Section("K", "Football field weights", "Extended", [
        Param("weight_dcf", "DCF weight", "percentage"),
        Param("weight_comps", "Comparable companies weight", "percentage"),
        Param("weight_precedent", "Precedent transactions weight", "percentage"),
        Param("weight_nav", "NAV / asset-based weight", "percentage"),
        Param("selected_low", "Selected valuation low", "currency"),
        Param("selected_mid", "Selected valuation mid", "currency"),
        Param("selected_high", "Selected valuation high", "currency"),
    ]),
    Section("L", "Sensitivity ranges", "Extended", [
        Param("sens_wacc_step", "WACC sensitivity step", "percentage"),
        Param("sens_wacc_count", "WACC steps each side of base", "number"),
        Param("sens_terminal_g_step", "Terminal g sensitivity step", "percentage"),
        Param("sens_terminal_g_count", "Terminal g steps each side", "number"),
        Param("sens_revenue_g_step", "Revenue growth sens step", "percentage"),
        Param("sens_ebitda_margin_step", "EBITDA margin sens step", "percentage"),
    ]),
]


# Tabular blocks (Sections G + H) live below the scalars
COCO_COLUMNS = [
    ("coco_tier", "Tier", "enum"),
    ("coco_include", "Include", "boolean"),
    ("coco_company", "Company", "text"),
    ("coco_ticker", "Ticker", "text"),
    ("coco_country", "Country", "enum"),
    ("coco_accounting", "Accounting standard", "enum"),
    ("coco_market_cap", "Market cap (USD mm)", "currency"),
    ("coco_d_to_e", "D/E ratio", "percentage"),
    ("coco_raw_beta", "Raw levered beta", "number"),
    ("coco_tax_rate", "Effective tax rate", "percentage"),
    ("coco_unlevered_beta", "Unlevered beta (calculated)", "number"),
]
COCO_ROWS = 30

PRECEDENT_COLUMNS = [
    ("precedent_include", "Include", "boolean"),
    ("precedent_date", "Date", "date"),
    ("precedent_acquirer", "Acquirer", "text"),
    ("precedent_target", "Target", "text"),
    ("precedent_ev", "EV (USD mm)", "currency"),
    ("precedent_ev_revenue", "EV/Revenue", "number"),
    ("precedent_ev_ebitda", "EV/EBITDA", "number"),
    ("precedent_premium", "Premium paid", "percentage"),
    ("precedent_rationale", "Strategic rationale", "text"),
]
PRECEDENT_ROWS = 15


# ---------------------------------------------------------------------------
# Sheet plan (23 sheets per v1 template)
# ---------------------------------------------------------------------------

SHEETS = [
    ("README", "Instructions, version, manual-input checklist"),
    ("Dashboard", "High-level outputs: EV, Net debt, Surplus, Equity, WACC, terminal g + benchmark charts"),
    ("Inputs", "Single source of truth — all driver parameters"),
    ("Historical FS", "3-5 yr audited income statement + balance sheet (PBC raw)"),
    ("Projections", "Year-by-year forecast lines driven by Inputs"),
    ("DCF", "FCFF schedule, partial-year discount, PV, terminal value"),
    ("DCF (Independent)", "Parallel scenario with independent WACC range"),
    ("Comps", "Apply CoCo multiples to target metrics; implied EV/Sales, EV/EBITDA, P/E"),
    ("Precedent", "Precedent M&A transactions and implied multiples"),
    ("Football Field", "Valuation range chart per methodology + weighted selected range"),
    ("WACC", "Full Ke + Kd build with Per-Mgmt and Independent scenarios"),
    ("Beta Analysis", "Comparable beta unlevering / relevering with regression detail"),
    ("CoCo Selection", "Comparable company selection with tier flags + market cap"),
    ("CoCo Multiples", "EV/Sales, EV/EBITDA, P/E per comparable"),
    ("CoCo Margins", "Gross / EBIT / net margin per comparable"),
    ("CoCo Ratios", "ROE / ROA / D/E / current ratio per comparable"),
    ("CIQ Data Timeline", "Capital IQ data import audit trail FY-4 to FY-1"),
    ("Country ERP", "Country risk premium reference table (Damodaran)"),
    ("Industry Averages", "Industry-level multiples + margins reference (Damodaran)"),
    ("Adjustments", "EV -> Equity bridge: net debt, surplus, DLOM, DLOC application"),
    ("Sensitivity", "WACC x terminal g; WACC x revenue growth; multiple x metric"),
    ("R&D + Lease Adj", "R&D capitalization + operating lease conversion (toggle-driven)"),
    ("Valuation Summary", "Final EV -> Equity -> Per share with all weighting and adjustments"),
]


# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------

THIN = Side(style="thin", color="888888")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
SECTION_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
EXTENDED_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
TODO_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FONT = Font(bold=True, size=11)
NORMAL_FONT = Font(size=10)
SMALL_FONT = Font(size=9, italic=True, color="555555")


# ---------------------------------------------------------------------------
# Build
# ---------------------------------------------------------------------------

SOURCE_ENUM = (
    '"Audited FS,Management Projections,Capital IQ,Bloomberg,Damodaran,'
    'Kroll,Mercer,Prospectus,Engagement Letter,Calculated,Manual"'
)
TYPE_HEADERS = ["id", "Parameter", "Type", "Value (Per Mgmt)", "Value (Independent)",
                "Source", "Source detail", "Notes"]


def write_header_band(ws, row: int, label: str, fill=HEADER_FILL, font=HEADER_FONT, span: int = 8):
    ws.cell(row=row, column=1, value=label)
    ws.cell(row=row, column=1).font = font
    ws.cell(row=row, column=1).fill = fill
    ws.cell(row=row, column=1).alignment = Alignment(vertical="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = 22


def build_inputs_sheet(ws, defined_names: list[tuple[str, str]]):
    """Build the Inputs sheet. Append (named_range, "'Sheet'!$X$Y") tuples to defined_names."""
    sheet_name = ws.title

    # Title band
    write_header_band(ws, 1, "Inputs — single source of truth for all valuation parameters")
    ws.cell(row=2, column=1, value=(
        "Every parameter below is a named range. Downstream sheets reference these by name, "
        "not by cell coordinate. The Source / Source detail / Notes columns record the audit trail."
    )).font = SMALL_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

    # Column header
    header_row = 4
    for i, h in enumerate(TYPE_HEADERS, 1):
        c = ws.cell(row=header_row, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER

    # Column widths
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["G"].width = 36
    ws.column_dimensions["H"].width = 30

    row = header_row + 2
    source_dv = DataValidation(type="list", formula1=SOURCE_ENUM, allow_blank=True)
    bool_dv = DataValidation(type="list", formula1='"true,false"', allow_blank=True)
    ws.add_data_validation(source_dv)
    ws.add_data_validation(bool_dv)

    # ----- Scalar sections -----
    for sect in SECTIONS:
        # Section banner
        fill = SECTION_FILL if sect.tier == "Core" else EXTENDED_FILL
        ws.cell(row=row, column=1, value=f"Section {sect.code} — {sect.title}  ({sect.tier})").font = SECTION_FONT
        ws.cell(row=row, column=1).fill = fill
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        ws.row_dimensions[row].height = 18
        row += 1

        for p in sect.params:
            ws.cell(row=row, column=1, value=p.pid).font = Font(name="Menlo", size=9, color="333333")
            ws.cell(row=row, column=2, value=p.label).font = NORMAL_FONT
            ws.cell(row=row, column=3, value=p.type_).font = SMALL_FONT
            # Value cells get a soft border + light background
            for col in (4, 5):
                cell = ws.cell(row=row, column=col)
                cell.border = BORDER
                cell.fill = INPUT_FILL
            ws.cell(row=row, column=6).border = BORDER
            ws.cell(row=row, column=7).border = BORDER
            ws.cell(row=row, column=8).border = BORDER

            # Apply data validation
            source_dv.add(f"F{row}")
            if p.type_ == "boolean":
                bool_dv.add(f"D{row}")
                if p.scenario:
                    bool_dv.add(f"E{row}")

            # Notes column gets the schema notes if provided
            if p.notes:
                ws.cell(row=row, column=8, value=p.notes).font = SMALL_FONT

            # Define named ranges
            if p.scenario:
                defined_names.append((
                    f"{p.pid}_per_mgmt",
                    f"'{sheet_name}'!$D${row}",
                ))
                defined_names.append((
                    f"{p.pid}_indep",
                    f"'{sheet_name}'!$E${row}",
                ))
            else:
                defined_names.append((
                    p.pid,
                    f"'{sheet_name}'!$D${row}",
                ))
            row += 1

        row += 1  # blank row between sections

    # ----- Section G — CoCo table -----
    ws.cell(row=row, column=1, value="Section G — Comparable companies table  (Core)").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    row += 1
    coco_header_row = row
    for i, (cid, clabel, ctype) in enumerate(COCO_COLUMNS, 1):
        c = ws.cell(row=row, column=i, value=clabel)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = BORDER
    row += 1
    coco_first_data_row = row
    for r in range(COCO_ROWS):
        for i in range(1, len(COCO_COLUMNS) + 1):
            ws.cell(row=row, column=i).border = BORDER
        row += 1
    coco_last_row = row - 1
    last_col_letter = get_column_letter(len(COCO_COLUMNS))
    defined_names.append((
        "cocos_table",
        f"'{sheet_name}'!$A${coco_first_data_row}:${last_col_letter}${coco_last_row}",
    ))
    row += 1

    # ----- Section H — Precedent table -----
    ws.cell(row=row, column=1, value="Section H — Precedent transactions  (Extended)").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = EXTENDED_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    row += 1
    for i, (cid, clabel, ctype) in enumerate(PRECEDENT_COLUMNS, 1):
        c = ws.cell(row=row, column=i, value=clabel)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = BORDER
    row += 1
    prec_first_data_row = row
    for r in range(PRECEDENT_ROWS):
        for i in range(1, len(PRECEDENT_COLUMNS) + 1):
            ws.cell(row=row, column=i).border = BORDER
        row += 1
    prec_last_row = row - 1
    last_col_letter_p = get_column_letter(len(PRECEDENT_COLUMNS))
    defined_names.append((
        "precedents_table",
        f"'{sheet_name}'!$A${prec_first_data_row}:${last_col_letter_p}${prec_last_row}",
    ))


PROJECTION_YEARS = 5  # explicit forecast horizon; matches sample_inputs.json


def col_for_year(y: int) -> str:
    """Year columns: Y0 (base) = C, Y1 = D, ..., Y5 = H."""
    return get_column_letter(3 + y)


def build_projections_formulas(ws):
    """Wire the Projections sheet. Reads Inputs named ranges. Y0 column is the
    base year (audited last twelve months) — for v1, user fills C8 (revenue) and
    C27 (NWC) manually. TODO: link to Historical FS once that sheet is built."""
    write_header_band(ws, 1, "Projections — driven by Inputs")
    ws.cell(row=2, column=1, value=(
        "Year-by-year forecast lines. Every assumption flows from Inputs named ranges. "
        "Edit a value on Inputs and this sheet recalculates."
    )).font = SMALL_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

    # Year header row
    header_row = 4
    ws.cell(row=header_row, column=1, value="Line item").font = HEADER_FONT
    ws.cell(row=header_row, column=2, value="Unit").font = HEADER_FONT
    ws.cell(row=header_row, column=3, value="Y0 (base)").font = HEADER_FONT
    for y in range(1, PROJECTION_YEARS + 1):
        ws.cell(row=header_row, column=3 + y, value=f"Y{y}").font = HEADER_FONT
    for col in range(1, 3 + PROJECTION_YEARS + 1):
        ws.cell(row=header_row, column=col).fill = HEADER_FILL
        ws.cell(row=header_row, column=col).alignment = Alignment(horizontal="center")
        ws.cell(row=header_row, column=col).border = BORDER
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 12
    for y in range(0, PROJECTION_YEARS + 1):
        ws.column_dimensions[col_for_year(y)].width = 16

    # Helper to lay down a section banner row
    def section(row: int, title: str) -> None:
        ws.cell(row=row, column=1, value=title).font = SECTION_FONT
        ws.cell(row=row, column=1).fill = SECTION_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)

    # Helper to write a labelled row
    def lrow(row: int, label: str, unit: str = "") -> None:
        ws.cell(row=row, column=1, value=label).font = NORMAL_FONT
        if unit:
            ws.cell(row=row, column=2, value=unit).font = SMALL_FONT

    # ----- Revenue & growth -----
    section(6, "Revenue & growth")
    R_REV, R_RG = 8, 9
    lrow(R_REV, "Revenue", "(currency × unit)")
    ws.cell(row=R_REV, column=3, value=None)  # base revenue manual fill (Y0)
    ws.cell(row=R_REV, column=3).fill = INPUT_FILL
    ws.cell(row=R_REV, column=3).border = BORDER
    ws.cell(row=R_REV, column=3).comment = None
    ws.cell(row=R_REV, column=2).font = SMALL_FONT
    # Mark Y0 cell as manual entry
    ws.cell(row=R_REV, column=2, value="(manual Y0)").font = SMALL_FONT
    for y in range(1, PROJECTION_YEARS + 1):
        prev = col_for_year(y - 1)
        cur = col_for_year(y)
        ws[f"{cur}{R_REV}"] = f"={prev}{R_REV}*(1+revenue_growth_y{y})"
        ws[f"{cur}{R_RG}"] = f"=revenue_growth_y{y}"
    lrow(R_RG, "Revenue growth", "%")

    # ----- Profitability -----
    section(11, "Profitability")
    R_GP, R_GM, R_OPEX, R_EBITDA, R_EBITDAM, R_DEP, R_EBIT, R_EBITM = 12, 13, 14, 15, 16, 17, 18, 19
    lrow(R_GP, "Gross profit", "(currency × unit)")
    lrow(R_GM, "Gross margin", "%")
    lrow(R_OPEX, "Operating expenses", "(currency × unit)")
    lrow(R_EBITDA, "EBITDA", "(currency × unit)")
    lrow(R_EBITDAM, "EBITDA margin", "%")
    lrow(R_DEP, "Depreciation & amortization", "(currency × unit)")
    lrow(R_EBIT, "EBIT", "(currency × unit)")
    lrow(R_EBITM, "EBIT margin", "%")
    for y in range(1, PROJECTION_YEARS + 1):
        cur = col_for_year(y)
        ws[f"{cur}{R_GP}"] = f"={cur}{R_REV}*gross_margin_y{y}"
        ws[f"{cur}{R_GM}"] = f"=gross_margin_y{y}"
        ws[f"{cur}{R_OPEX}"] = f"=-{cur}{R_REV}*opex_pct_revenue_y{y}"
        ws[f"{cur}{R_EBITDA}"] = f"={cur}{R_GP}+{cur}{R_OPEX}"
        ws[f"{cur}{R_EBITDAM}"] = f"=IFERROR({cur}{R_EBITDA}/{cur}{R_REV},0)"
        ws[f"{cur}{R_DEP}"] = f"=-{cur}{R_REV}*dep_pct_revenue_y{y}"
        ws[f"{cur}{R_EBIT}"] = f"={cur}{R_EBITDA}+{cur}{R_DEP}"
        ws[f"{cur}{R_EBITM}"] = f"=IFERROR({cur}{R_EBIT}/{cur}{R_REV},0)"

    # ----- Tax & net income -----
    section(21, "Tax & net income")
    R_TAX, R_NI = 22, 23
    lrow(R_TAX, "Tax", "(currency × unit)")
    lrow(R_NI, "Net income", "(currency × unit)")
    for y in range(1, PROJECTION_YEARS + 1):
        cur = col_for_year(y)
        # Branching: explicit override > two_tier > flat (use tax_rate_high as flat rate)
        # MAX(EBIT,0) suppresses negative-tax (we don't model loss carryforward in v1)
        formula = (
            f'=IF(ISNUMBER(tax_effective_rate),'
            f'-MAX({cur}{R_EBIT},0)*tax_effective_rate,'
            f'IF(EXACT(tax_type,"two_tier"),'
            f'IF({cur}{R_EBIT}<=tax_threshold,'
            f'-MAX({cur}{R_EBIT},0)*tax_rate_low,'
            f'-(tax_threshold*tax_rate_low+MAX({cur}{R_EBIT}-tax_threshold,0)*tax_rate_high)),'
            f'-MAX({cur}{R_EBIT},0)*tax_rate_high))'
        )
        ws[f"{cur}{R_TAX}"] = formula
        ws[f"{cur}{R_NI}"] = f"={cur}{R_EBIT}+{cur}{R_TAX}"

    # ----- Capex & working capital -----
    section(25, "Capex & working capital")
    R_CAPEX, R_NWC, R_DNWC = 26, 27, 28
    lrow(R_CAPEX, "Capex", "(currency × unit)")
    lrow(R_NWC, "Net working capital", "(currency × unit)")
    lrow(R_DNWC, "Δ Working capital", "(currency × unit)")
    # Y0 NWC = manual entry (needed as base for ΔNWC in Y1)
    ws.cell(row=R_NWC, column=3).fill = INPUT_FILL
    ws.cell(row=R_NWC, column=3).border = BORDER
    ws.cell(row=R_NWC, column=2, value="(manual Y0)").font = SMALL_FONT
    for y in range(1, PROJECTION_YEARS + 1):
        cur = col_for_year(y)
        prev = col_for_year(y - 1)
        ws[f"{cur}{R_CAPEX}"] = f"=-{cur}{R_REV}*capex_pct_revenue_y{y}"
        ws[f"{cur}{R_NWC}"] = f"={cur}{R_REV}*nwc_pct_sales_y{y}"
        ws[f"{cur}{R_DNWC}"] = f"=-({cur}{R_NWC}-{prev}{R_NWC})"

    # ----- FCFF -----
    section(30, "FCFF construction")
    R_F_EBIT, R_F_TAX, R_F_DEP, R_F_CAPEX, R_F_DWC, R_FCFF = 31, 32, 33, 34, 35, 36
    lrow(R_F_EBIT, "EBIT", "(currency × unit)")
    lrow(R_F_TAX, "Less: Tax on EBIT", "(currency × unit)")
    lrow(R_F_DEP, "Add: Depreciation", "(currency × unit)")
    lrow(R_F_CAPEX, "Less: Capex", "(currency × unit)")
    lrow(R_F_DWC, "Less: Δ Working capital", "(currency × unit)")
    lrow(R_FCFF, "Free Cash Flow to Firm (FCFF)", "(currency × unit)")
    for y in range(1, PROJECTION_YEARS + 1):
        cur = col_for_year(y)
        ws[f"{cur}{R_F_EBIT}"] = f"={cur}{R_EBIT}"
        ws[f"{cur}{R_F_TAX}"] = f"={cur}{R_TAX}"  # already negative
        ws[f"{cur}{R_F_DEP}"] = f"=-{cur}{R_DEP}"  # flip stored negative back to positive add-back
        ws[f"{cur}{R_F_CAPEX}"] = f"={cur}{R_CAPEX}"  # already negative
        ws[f"{cur}{R_F_DWC}"] = f"={cur}{R_DNWC}"  # already signed
        ws[f"{cur}{R_FCFF}"] = f"=SUM({cur}{R_F_EBIT}:{cur}{R_F_DWC})"
        # Bold the FCFF row
        ws.cell(row=R_FCFF, column=3 + y).font = Font(bold=True)
    ws.cell(row=R_FCFF, column=1).font = Font(bold=True)

    # Format hints — apply percentage format to margin rows; integer to monetary
    pct_rows = [R_RG, R_GM, R_EBITDAM, R_EBITM]
    for r in pct_rows:
        for y in range(1, PROJECTION_YEARS + 1):
            ws.cell(row=r, column=3 + y).number_format = "0.0%"
    money_rows = [R_REV, R_GP, R_OPEX, R_EBITDA, R_DEP, R_EBIT, R_TAX, R_NI,
                  R_CAPEX, R_NWC, R_DNWC, R_F_EBIT, R_F_TAX, R_F_DEP, R_F_CAPEX, R_F_DWC, R_FCFF]
    for r in money_rows:
        for y in range(0, PROJECTION_YEARS + 1):
            ws.cell(row=r, column=3 + y).number_format = "#,##0"


def build_wacc_formulas(ws):
    """Wire the WACC sheet. Two scenarios side by side: Per Management (col E) and
    Independent (col F). Builds Ke -> Kd -> WACC; outputs cycle back to Inputs
    sheet's calculated cells via post-pass."""
    write_header_band(ws, 1, "WACC — discount rate build-up")
    ws.cell(row=2, column=1, value=(
        "Two scenarios. Inputs (Rf, ERP, country premium) shared; beta / D/E / size / "
        "specific / Kd are scenario-specific. Calculated outputs (levered beta, Ke, "
        "after-tax Kd, WACC) cycle back to the Inputs sheet."
    )).font = SMALL_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)

    # Header row
    HDR = 4
    headers = ["Component", "", "Type", "", "Per Management", "Independent"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=HDR, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 4
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 4
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18

    def section(row, title):
        ws.cell(row=row, column=1, value=title).font = SECTION_FONT
        ws.cell(row=row, column=1).fill = SECTION_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)

    def labelled(row, label, type_="%"):
        ws.cell(row=row, column=1, value=label).font = NORMAL_FONT
        ws.cell(row=row, column=3, value=type_).font = SMALL_FONT
        for col in (5, 6):
            ws.cell(row=row, column=col).border = BORDER
            ws.cell(row=row, column=col).number_format = "0.00%" if type_ == "%" else "0.000"

    # Effective tax rate proxy used for unlevering / Kd after-tax
    section(6, "Cost of Equity (Ke)")
    R_T, R_UB, R_DE, R_LB = 7, 8, 9, 10
    labelled(R_T, "Effective tax rate", "%")
    labelled(R_UB, "Unlevered beta", "x")
    labelled(R_DE, "D/E ratio", "%")
    labelled(R_LB, "Levered beta (calc)", "x")
    R_RF, R_ERP, R_CRP, R_SP, R_SR, R_KE = 11, 12, 13, 14, 15, 16
    labelled(R_RF, "Risk-free rate", "%")
    labelled(R_ERP, "Equity risk premium (ERP)", "%")
    labelled(R_CRP, "Country risk premium", "%")
    labelled(R_SP, "Size premium", "%")
    labelled(R_SR, "Specific risk premium", "%")
    labelled(R_KE, "Cost of Equity (calc)", "%")
    ws.cell(row=R_KE, column=1).font = Font(bold=True)

    section(18, "Cost of Debt (Kd)")
    R_KDP, R_KDA = 19, 20
    labelled(R_KDP, "Pre-tax cost of debt", "%")
    labelled(R_KDA, "After-tax cost of debt (calc)", "%")
    ws.cell(row=R_KDA, column=1).font = Font(bold=True)

    section(22, "Capital structure & WACC")
    R_DV, R_EV, R_W = 23, 24, 25
    labelled(R_DV, "D/V (debt weight)", "%")
    labelled(R_EV, "E/V (equity weight)", "%")
    labelled(R_W, "WACC (calc)", "%")
    ws.cell(row=R_W, column=1).font = Font(bold=True)

    # Effective tax rate fallback: explicit override > tax_rate_high (as flat)
    tax_formula = '=IF(ISNUMBER(tax_effective_rate),tax_effective_rate,tax_rate_high)'

    # Per-Management column (E) and Independent column (F)
    for col_letter, suf in (("E", "per_mgmt"), ("F", "indep")):
        ws[f"{col_letter}{R_T}"] = tax_formula
        ws[f"{col_letter}{R_UB}"] = f"=unlevered_beta_{suf}"
        ws[f"{col_letter}{R_DE}"] = f"=target_debt_to_equity_{suf}"
        ws[f"{col_letter}{R_LB}"] = (
            f"={col_letter}{R_UB}*(1+(1-{col_letter}{R_T})*{col_letter}{R_DE})"
        )
        ws[f"{col_letter}{R_RF}"] = "=risk_free_rate"
        ws[f"{col_letter}{R_ERP}"] = "=equity_risk_premium"
        ws[f"{col_letter}{R_CRP}"] = "=country_risk_premium"
        ws[f"{col_letter}{R_SP}"] = f"=size_premium_{suf}"
        ws[f"{col_letter}{R_SR}"] = f"=specific_risk_premium_{suf}"
        ws[f"{col_letter}{R_KE}"] = (
            f"={col_letter}{R_RF}+{col_letter}{R_LB}*{col_letter}{R_ERP}"
            f"+{col_letter}{R_CRP}+{col_letter}{R_SP}+{col_letter}{R_SR}"
        )
        ws[f"{col_letter}{R_KDP}"] = f"=pretax_cost_of_debt_{suf}"
        ws[f"{col_letter}{R_KDA}"] = f"={col_letter}{R_KDP}*(1-{col_letter}{R_T})"
        ws[f"{col_letter}{R_DV}"] = f"=target_debt_weight_{suf}"
        ws[f"{col_letter}{R_EV}"] = f"=target_equity_weight_{suf}"
        ws[f"{col_letter}{R_W}"] = (
            f"={col_letter}{R_KE}*{col_letter}{R_EV}+{col_letter}{R_KDA}*{col_letter}{R_DV}"
        )

    # Stash output cell addresses on the worksheet for the post-pass to find
    ws._wacc_output_cells = {  # type: ignore[attr-defined]
        "levered_beta": (R_LB, "E", "F"),
        "cost_of_equity": (R_KE, "E", "F"),
        "aftertax_cost_of_debt": (R_KDA, "E", "F"),
        "wacc": (R_W, "E", "F"),
    }


def wire_calculated_inputs(wb):
    """Post-pass: the Inputs sheet has rows for calculated params (levered_beta,
    cost_of_equity, aftertax_cost_of_debt, wacc) — fill those cells with formulas
    that pull from the WACC sheet, so named ranges resolve to live values."""
    if "WACC" not in wb.sheetnames or "Inputs" not in wb.sheetnames:
        return
    wacc_ws = wb["WACC"]
    inputs_ws = wb["Inputs"]
    output_cells = getattr(wacc_ws, "_wacc_output_cells", {})
    if not output_cells:
        return
    for stem, (wacc_row, pm_col, indep_col) in output_cells.items():
        for suffix, wacc_col in (("per_mgmt", pm_col), ("indep", indep_col)):
            named = f"{stem}_{suffix}"
            if named not in wb.defined_names:
                continue
            text = wb.defined_names[named].attr_text
            if "!" not in text or ":" in text:
                continue
            sheet_part, addr = text.rsplit("!", 1)
            sheet_part = sheet_part.strip("'")
            addr = addr.replace("$", "")
            ws = wb[sheet_part]
            ws[addr] = f"=WACC!{wacc_col}{wacc_row}"
            # Format as percentage / number
            ws[addr].number_format = "0.00%" if stem != "levered_beta" else "0.000"


def build_dcf_formulas(ws, scenario: str = "per_mgmt"):
    """Wire a DCF sheet for one scenario (per_mgmt or indep). Reads FCFF from
    Projections row 36, discounts at the scenario WACC, applies Gordon Growth
    or Exit Multiple terminal value, and sums to Enterprise Value.

    Convention: Y_n discount period = n (no stub year in v1). Terminal value
    is at end of Y5; PV'd back at the same Y5 discount factor.
    """
    title_suffix = "Per Management" if scenario == "per_mgmt" else "Independent"
    write_header_band(ws, 1, f"DCF — {title_suffix} scenario")
    ws.cell(row=2, column=1, value=(
        "FCFF schedule from Projections, discounted at scenario WACC. "
        "Convention: full-period discounting (no stub-year partial); terminal "
        "value computed at end of Y5 and PV'd at the same Y5 discount factor."
    )).font = SMALL_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)

    # Header row — D..H = Y1..Y5, I = Terminal
    HDR = 4
    ws.cell(row=HDR, column=1, value="Item").font = HEADER_FONT
    ws.cell(row=HDR, column=2, value="Unit").font = HEADER_FONT
    for y in range(1, PROJECTION_YEARS + 1):
        ws.cell(row=HDR, column=3 + y, value=f"Y{y}").font = HEADER_FONT
    ws.cell(row=HDR, column=3 + PROJECTION_YEARS + 1, value="Terminal").font = HEADER_FONT
    for col in range(1, 3 + PROJECTION_YEARS + 2):
        ws.cell(row=HDR, column=col).fill = HEADER_FILL
        ws.cell(row=HDR, column=col).alignment = Alignment(horizontal="center")
        ws.cell(row=HDR, column=col).border = BORDER
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 12
    for c in range(3, 3 + PROJECTION_YEARS + 2):
        ws.column_dimensions[get_column_letter(c)].width = 14

    def section(row, title):
        ws.cell(row=row, column=1, value=title).font = SECTION_FONT
        ws.cell(row=row, column=1).fill = SECTION_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)

    def lrow(row, label, unit=""):
        ws.cell(row=row, column=1, value=label).font = NORMAL_FONT
        if unit:
            ws.cell(row=row, column=2, value=unit).font = SMALL_FONT

    wacc_ref = f"wacc_{scenario}"

    # ----- Inputs banner -----
    section(6, "Discount inputs")
    R_WACC, R_G = 7, 8
    lrow(R_WACC, "WACC (this scenario)", "%")
    lrow(R_G, "Terminal growth rate", "%")
    ws.cell(row=R_WACC, column=4, value=f"={wacc_ref}").number_format = "0.00%"
    ws.cell(row=R_G, column=4, value="=terminal_growth_rate").number_format = "0.00%"

    # ----- FCFF & discount -----
    section(10, "FCFF & discount mechanics")
    R_FCFF, R_PERIOD, R_DF, R_PV = 11, 12, 13, 14
    lrow(R_FCFF, "FCFF", "(currency × unit)")
    lrow(R_PERIOD, "Discount period", "years")
    lrow(R_DF, "Discount factor", "x")
    lrow(R_PV, "PV of FCFF", "(currency × unit)")
    for y in range(1, PROJECTION_YEARS + 1):
        col = get_column_letter(3 + y)
        ws[f"{col}{R_FCFF}"] = f"=Projections!{col}36"  # Projections row 36 = FCFF
        ws[f"{col}{R_PERIOD}"] = y
        ws[f"{col}{R_DF}"] = f"=1/(1+{wacc_ref})^{col}{R_PERIOD}"
        ws[f"{col}{R_PV}"] = f"={col}{R_FCFF}*{col}{R_DF}"
        ws[f"{col}{R_FCFF}"].number_format = "#,##0"
        ws[f"{col}{R_PERIOD}"].number_format = "0.00"
        ws[f"{col}{R_DF}"].number_format = "0.0000"
        ws[f"{col}{R_PV}"].number_format = "#,##0"

    # ----- Terminal value -----
    section(16, "Terminal value")
    R_TM, R_TFCFF, R_TGORDON, R_TEXIT, R_TSEL, R_TPV = 17, 18, 19, 20, 21, 22
    lrow(R_TM, "Method", "")
    lrow(R_TFCFF, "Y5 FCFF", "(currency × unit)")
    lrow(R_TGORDON, "Terminal value — Gordon Growth", "(currency × unit)")
    lrow(R_TEXIT, "Terminal value — Exit Multiple", "(currency × unit)")
    lrow(R_TSEL, "Selected terminal value", "(currency × unit)")
    lrow(R_TPV, "PV of terminal value", "(currency × unit)")
    # Use column I (year+1 column)
    TC = get_column_letter(3 + PROJECTION_YEARS + 1)  # I
    LAST_Y_COL = get_column_letter(3 + PROJECTION_YEARS)  # H
    ws[f"{TC}{R_TM}"] = "=terminal_method"
    ws[f"{TC}{R_TFCFF}"] = f"=Projections!{LAST_Y_COL}36"
    # Gordon: FCFF_y5 × (1+g) / (WACC - g)
    ws[f"{TC}{R_TGORDON}"] = (
        f"=IFERROR({TC}{R_TFCFF}*(1+terminal_growth_rate)/({wacc_ref}-terminal_growth_rate),0)"
    )
    # Exit multiple — switches on type: EV/EBITDA -> Y5 EBITDA × mult, EV/Sales -> Y5 Revenue × mult, P/E -> Y5 NI × mult
    ws[f"{TC}{R_TEXIT}"] = (
        f'=IF(EXACT(terminal_exit_multiple_type,"EV/EBITDA"),'
        f'Projections!{LAST_Y_COL}15*terminal_exit_multiple_value,'
        f'IF(EXACT(terminal_exit_multiple_type,"EV/Sales"),'
        f'Projections!{LAST_Y_COL}8*terminal_exit_multiple_value,'
        f'IF(EXACT(terminal_exit_multiple_type,"P/E"),'
        f'Projections!{LAST_Y_COL}23*terminal_exit_multiple_value,0)))'
    )
    ws[f"{TC}{R_TSEL}"] = (
        f'=IF(EXACT(terminal_method,"exit_multiple"),{TC}{R_TEXIT},{TC}{R_TGORDON})'
    )
    # PV of terminal value: discount at Y5's discount factor
    ws[f"{TC}{R_TPV}"] = f"={TC}{R_TSEL}*{LAST_Y_COL}{R_DF}"
    for r in (R_TFCFF, R_TGORDON, R_TEXIT, R_TSEL, R_TPV):
        ws.cell(row=r, column=3 + PROJECTION_YEARS + 1).number_format = "#,##0"

    # ----- Enterprise Value -----
    section(24, "Enterprise Value")
    R_SUMPV, R_TPV2, R_EV = 25, 26, 27
    lrow(R_SUMPV, "Sum of PV of explicit-period FCFF")
    lrow(R_TPV2, "+ PV of terminal value")
    lrow(R_EV, "Enterprise Value")
    first_year_col = get_column_letter(4)
    ws[f"{TC}{R_SUMPV}"] = f"=SUM({first_year_col}{R_PV}:{LAST_Y_COL}{R_PV})"
    ws[f"{TC}{R_TPV2}"] = f"={TC}{R_TPV}"
    ws[f"{TC}{R_EV}"] = f"={TC}{R_SUMPV}+{TC}{R_TPV2}"
    for r in (R_SUMPV, R_TPV2, R_EV):
        cell = ws.cell(row=r, column=3 + PROJECTION_YEARS + 1)
        cell.number_format = "#,##0"
        if r == R_EV:
            cell.font = Font(bold=True)
    ws.cell(row=R_EV, column=1).font = Font(bold=True)

    # Stash the EV cell address for downstream sheets
    ws._dcf_ev_cell = f"{TC}{R_EV}"  # type: ignore[attr-defined]


def build_adjustments_formulas(ws):
    """EV -> Equity bridge for both scenarios. Pulls EV from DCF and DCF (Independent),
    applies surplus / non-operating / net debt / minority interests, then DLOM and DLOC."""
    write_header_band(ws, 1, "Adjustments — EV → Equity bridge")
    ws.cell(row=2, column=1, value=(
        "Applies surplus assets, net debt, minority interests, then DLOM and DLOC, "
        "to derive Fair Value of the client's equity interest. Both scenarios computed "
        "side by side."
    )).font = SMALL_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)

    HDR = 4
    headers = ["Item", "", "Per Management", "Independent", "Notes"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=HDR, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 4
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 30

    def section(row, title):
        ws.cell(row=row, column=1, value=title).font = SECTION_FONT
        ws.cell(row=row, column=1).fill = SECTION_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)

    def lrow(row, label, fmt="#,##0"):
        ws.cell(row=row, column=1, value=label).font = NORMAL_FONT
        for col in (3, 4):
            ws.cell(row=row, column=col).number_format = fmt
            ws.cell(row=row, column=col).border = BORDER

    # ----- Bridge -----
    section(6, "EV → Equity bridge")
    R_EV, R_SURP, R_NOA, R_ND, R_MI, R_EQ_PRE = 7, 8, 9, 10, 11, 12
    lrow(R_EV, "Enterprise Value (from DCF)")
    lrow(R_SURP, "+ Surplus assets")
    lrow(R_NOA, "+ Non-operating assets")
    lrow(R_ND, "− Net debt")
    lrow(R_MI, "− Minority interests")
    lrow(R_EQ_PRE, "Equity value (pre-discounts)")
    ws.cell(row=R_EQ_PRE, column=1).font = Font(bold=True)

    # Per Mgmt = col C, Independent = col D
    PM, IN = "C", "D"
    ws[f"{PM}{R_EV}"] = "='DCF'!I27"
    ws[f"{IN}{R_EV}"] = "='DCF (Independent)'!I27"
    for col_letter in (PM, IN):
        ws[f"{col_letter}{R_SURP}"] = "=surplus_assets"
        ws[f"{col_letter}{R_NOA}"] = "=non_operating_assets"
        ws[f"{col_letter}{R_ND}"] = (
            "=-IF(ISNUMBER(net_debt_override),net_debt_override,0)"
        )
        ws[f"{col_letter}{R_MI}"] = "=-minority_interests"
        ws[f"{col_letter}{R_EQ_PRE}"] = (
            f"={col_letter}{R_EV}+{col_letter}{R_SURP}+{col_letter}{R_NOA}"
            f"+{col_letter}{R_ND}+{col_letter}{R_MI}"
        )
        ws.cell(row=R_EQ_PRE, column=ord(col_letter) - ord("A") + 1).font = Font(bold=True)

    # ----- DLOM / DLOC -----
    section(14, "Marketability & control discounts")
    R_DLOM_RATE, R_DLOM_AMT, R_EQ_DLOM, R_DLOC_RATE, R_DLOC_AMT, R_EQ_DLOC = 15, 16, 17, 18, 19, 20
    lrow(R_DLOM_RATE, "DLOM rate", "0.0%")
    lrow(R_DLOM_AMT, "Less: DLOM")
    lrow(R_EQ_DLOM, "Equity value after DLOM")
    lrow(R_DLOC_RATE, "DLOC rate", "0.0%")
    lrow(R_DLOC_AMT, "Less: DLOC")
    lrow(R_EQ_DLOC, "Equity value after DLOM and DLOC")
    ws.cell(row=R_EQ_DLOC, column=1).font = Font(bold=True)
    for col_letter in (PM, IN):
        ws[f"{col_letter}{R_DLOM_RATE}"] = "=dlom_pct"
        ws[f"{col_letter}{R_DLOM_AMT}"] = f"=-{col_letter}{R_EQ_PRE}*{col_letter}{R_DLOM_RATE}"
        ws[f"{col_letter}{R_EQ_DLOM}"] = f"={col_letter}{R_EQ_PRE}+{col_letter}{R_DLOM_AMT}"
        ws[f"{col_letter}{R_DLOC_RATE}"] = "=dloc_pct"
        ws[f"{col_letter}{R_DLOC_AMT}"] = f"=-{col_letter}{R_EQ_DLOM}*{col_letter}{R_DLOC_RATE}"
        ws[f"{col_letter}{R_EQ_DLOC}"] = f"={col_letter}{R_EQ_DLOM}+{col_letter}{R_DLOC_AMT}"
        ws.cell(row=R_EQ_DLOC, column=ord(col_letter) - ord("A") + 1).font = Font(bold=True)

    # ----- Client interest -----
    section(22, "Client's equity interest")
    R_INT_RATE, R_INT_VALUE = 23, 24
    lrow(R_INT_RATE, "Equity interest held by client", "0.0%")
    lrow(R_INT_VALUE, "Fair value of client's interest")
    ws.cell(row=R_INT_VALUE, column=1).font = Font(bold=True)
    for col_letter in (PM, IN):
        ws[f"{col_letter}{R_INT_RATE}"] = "=equity_interest_pct"
        ws[f"{col_letter}{R_INT_VALUE}"] = f"={col_letter}{R_EQ_DLOC}*{col_letter}{R_INT_RATE}"
        ws.cell(row=R_INT_VALUE, column=ord(col_letter) - ord("A") + 1).font = Font(bold=True)

    # Stash row addresses for Valuation Summary
    ws._adj_rows = {  # type: ignore[attr-defined]
        "ev": R_EV,
        "eq_pre": R_EQ_PRE,
        "eq_dlom": R_EQ_DLOM,
        "eq_dloc": R_EQ_DLOC,
        "interest_value": R_INT_VALUE,
    }


def build_valuation_summary_formulas(ws):
    """Headline outputs: EV, Equity, Per-share. Reads from Adjustments sheet for both
    scenarios. Per-share uses unit-aware conversion to actual currency."""
    write_header_band(ws, 1, "Valuation Summary")
    ws.cell(row=2, column=1, value=(
        "Headline outputs across both scenarios. Per-share values convert from "
        "(currency × unit) to actual currency using the unit selected on Inputs."
    )).font = SMALL_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)

    HDR = 4
    headers = ["Metric", "Unit", "Per Management", "Independent", "Average"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=HDR, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    def section(row, title):
        ws.cell(row=row, column=1, value=title).font = SECTION_FONT
        ws.cell(row=row, column=1).fill = SECTION_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)

    def lrow(row, label, unit="(currency × unit)", fmt="#,##0"):
        ws.cell(row=row, column=1, value=label).font = NORMAL_FONT
        ws.cell(row=row, column=2, value=unit).font = SMALL_FONT
        for col in (3, 4, 5):
            ws.cell(row=row, column=col).number_format = fmt
            ws.cell(row=row, column=col).border = BORDER

    # Reference the Adjustments rows by hardcoding the row numbers we just used
    # (matches build_adjustments_formulas above)
    A_EV = 7
    A_EQ_PRE = 12
    A_EQ_DLOM = 17
    A_EQ_DLOC = 20
    A_INT = 24
    PM_COL = "C"
    IN_COL = "D"

    section(6, "Headline values")
    R = 7
    headline = [
        ("Enterprise Value", A_EV),
        ("Equity value (pre-discounts)", A_EQ_PRE),
        ("Equity value after DLOM", A_EQ_DLOM),
        ("Equity value after DLOM and DLOC", A_EQ_DLOC),
        ("Fair value of client's interest", A_INT),
    ]
    for label, adj_row in headline:
        lrow(R, label)
        ws[f"C{R}"] = f"=Adjustments!{PM_COL}{adj_row}"
        ws[f"D{R}"] = f"=Adjustments!{IN_COL}{adj_row}"
        ws[f"E{R}"] = f"=AVERAGE(C{R}:D{R})"
        if "Fair value" in label:
            for col in ("C", "D", "E"):
                ws[f"{col}{R}"].font = Font(bold=True)
            ws.cell(row=R, column=1).font = Font(bold=True)
        R += 1

    # Per-share
    section(R + 1, "Per-share value")
    R += 2

    # Unit multiplier helper formula — '000 -> 1000, million -> 1e6, actual -> 1
    unit_mult = (
        '=IF(EXACT(unit,"\'000"),1000,'
        'IF(EXACT(unit,"million"),1000000,1))'
    )
    R_UNIT_MULT = R
    lrow(R, "Unit multiplier (to actual currency)", "x", "#,##0")
    ws[f"C{R}"] = unit_mult
    ws[f"D{R}"] = unit_mult
    ws[f"E{R}"] = ""
    R += 1

    R_SHARES_BASIC = R
    lrow(R, "Shares outstanding (basic)", "shares", "#,##0")
    ws[f"C{R}"] = "=shares_outstanding"
    ws[f"D{R}"] = "=shares_outstanding"
    R += 1

    R_PS_BASIC = R
    lrow(R, "Equity value per share (basic)", "actual currency", "#,##0.00")
    ws[f"C{R}"] = (
        f"=IFERROR(Adjustments!{PM_COL}{A_EQ_DLOC}*C{R_UNIT_MULT}/C{R_SHARES_BASIC},0)"
    )
    ws[f"D{R}"] = (
        f"=IFERROR(Adjustments!{IN_COL}{A_EQ_DLOC}*D{R_UNIT_MULT}/D{R_SHARES_BASIC},0)"
    )
    ws[f"E{R}"] = f"=AVERAGE(C{R}:D{R})"
    for col in ("C", "D", "E"):
        ws[f"{col}{R}"].font = Font(bold=True)
    ws.cell(row=R, column=1).font = Font(bold=True)
    R += 1

    R_SHARES_DIL = R
    lrow(R, "Shares outstanding (diluted)", "shares", "#,##0")
    ws[f"C{R}"] = "=shares_outstanding_diluted"
    ws[f"D{R}"] = "=shares_outstanding_diluted"
    R += 1

    R_PS_DIL = R
    lrow(R, "Equity value per share (diluted)", "actual currency", "#,##0.00")
    ws[f"C{R}"] = (
        f"=IFERROR(Adjustments!{PM_COL}{A_EQ_DLOC}*C{R_UNIT_MULT}/C{R_SHARES_DIL},0)"
    )
    ws[f"D{R}"] = (
        f"=IFERROR(Adjustments!{IN_COL}{A_EQ_DLOC}*D{R_UNIT_MULT}/D{R_SHARES_DIL},0)"
    )
    ws[f"E{R}"] = f"=AVERAGE(C{R}:D{R})"

    # Discount rates
    R += 2
    section(R, "Discount rate inputs (reference)")
    R += 1
    lrow(R, "WACC", "%", "0.00%")
    ws[f"C{R}"] = "=wacc_per_mgmt"
    ws[f"D{R}"] = "=wacc_indep"
    R += 1
    lrow(R, "Terminal growth rate", "%", "0.00%")
    ws[f"C{R}"] = "=terminal_growth_rate"
    ws[f"D{R}"] = "=terminal_growth_rate"


COCO_TABLE_FIRST_ROW = 127  # cocos_table starts at Inputs!A127 (must match build_inputs_sheet)
COCO_TABLE_ROWS = COCO_ROWS  # 30


def build_coco_selection_formulas(ws):
    """CoCo Selection — full mirror view of the cocos_table on Inputs."""
    write_header_band(ws, 1, "CoCo Selection — comparable companies")
    ws.cell(row=2, column=1, value=(
        "Mirror view of the comparable companies table on the Inputs sheet. "
        "Edit values on Inputs; this sheet refreshes automatically."
    )).font = SMALL_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=11)

    HDR = 4
    for i, (cid, clabel, ctype) in enumerate(COCO_COLUMNS, 1):
        c = ws.cell(row=HDR, column=i, value=clabel)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = BORDER
    for col in range(1, len(COCO_COLUMNS) + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = 14
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 14

    # Data rows mirroring Inputs!A127:K156
    first_data_row = HDR + 2
    for i in range(COCO_TABLE_ROWS):
        target_row = first_data_row + i
        source_row = COCO_TABLE_FIRST_ROW + i
        for col_idx in range(1, len(COCO_COLUMNS) + 1):
            col_letter = get_column_letter(col_idx)
            ws.cell(row=target_row, column=col_idx,
                    value=f"=Inputs!{col_letter}{source_row}")
            ws.cell(row=target_row, column=col_idx).border = BORDER

    # Compute unlevered beta on the LAST column (col 11) for each CoCo:
    # unlevered = raw_beta / (1 + (1 - tax_rate) * D/E)
    # Override the simple mirror with the calc since col K on Inputs is blank
    for i in range(COCO_TABLE_ROWS):
        target_row = first_data_row + i
        source_row = COCO_TABLE_FIRST_ROW + i
        ws.cell(row=target_row, column=11, value=(
            f"=IFERROR(Inputs!I{source_row}/"
            f"(1+(1-Inputs!J{source_row})*Inputs!H{source_row}),\"\")"
        ))


def build_coco_metric_sheet(ws, title: str, metric_columns: list[tuple[str, str]]):
    """Generic builder for CoCo Multiples / Margins / Ratios sheets.

    Each sheet mirrors Tier / Include / Company / Ticker from Inputs cocos_table,
    then has user-input columns for the metric. Below the table: Min / Q1 / Median
    / Mean / Q3 / Max + tier-filtered means.

    metric_columns: list of (label, number_format) tuples.
    """
    write_header_band(ws, 1, title)
    ws.cell(row=2, column=1, value=(
        "Per-CoCo metric values (Capital IQ / Bloomberg). Cols A-D mirror Inputs. "
        "Cols E onward are manual data entry. Summary stats compute below."
    )).font = SMALL_FONT
    n_total_cols = 4 + len(metric_columns)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_total_cols)

    # Header
    HDR = 4
    base_headers = ["Tier", "Include", "Company", "Ticker"]
    for i, h in enumerate(base_headers, 1):
        c = ws.cell(row=HDR, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = BORDER
    for j, (label, fmt) in enumerate(metric_columns):
        col = 5 + j
        c = ws.cell(row=HDR, column=col, value=label)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 14

    # Mirror rows for cocos
    first_data_row = HDR + 2
    last_data_row = first_data_row + COCO_TABLE_ROWS - 1
    for i in range(COCO_TABLE_ROWS):
        target_row = first_data_row + i
        source_row = COCO_TABLE_FIRST_ROW + i
        ws.cell(row=target_row, column=1, value=f"=Inputs!A{source_row}")  # tier
        ws.cell(row=target_row, column=2, value=f"=Inputs!B{source_row}")  # include
        ws.cell(row=target_row, column=3, value=f"=Inputs!C{source_row}")  # company
        ws.cell(row=target_row, column=4, value=f"=Inputs!D{source_row}")  # ticker
        for col in range(1, n_total_cols + 1):
            ws.cell(row=target_row, column=col).border = BORDER
            if col >= 5:
                # Apply metric format to data input cells
                fmt = metric_columns[col - 5][1]
                ws.cell(row=target_row, column=col).number_format = fmt
                ws.cell(row=target_row, column=col).fill = INPUT_FILL

    # Summary statistics
    stats_start = last_data_row + 2
    stats_rows = [
        ("Maximum", "MAX"),
        ("Q3 (75%)", "QUARTILE_3"),
        ("Median", "MEDIAN"),
        ("Mean", "AVERAGE"),
        ("Q1 (25%)", "QUARTILE_1"),
        ("Minimum", "MIN"),
    ]
    for i, (label, fn) in enumerate(stats_rows):
        r = stats_start + i
        ws.cell(row=r, column=4, value=label).font = Font(bold=True)
        for j, (mlabel, fmt) in enumerate(metric_columns):
            col = 5 + j
            data_range = f"{get_column_letter(col)}{first_data_row}:{get_column_letter(col)}{last_data_row}"
            if fn == "QUARTILE_3":
                f = f"=IFERROR(QUARTILE({data_range},3),\"n/a\")"
            elif fn == "QUARTILE_1":
                f = f"=IFERROR(QUARTILE({data_range},1),\"n/a\")"
            else:
                f = f"=IFERROR({fn}({data_range}),\"n/a\")"
            ws.cell(row=r, column=col, value=f)
            ws.cell(row=r, column=col).number_format = fmt
            ws.cell(row=r, column=col).font = Font(bold=(label == "Median"))

    # Tier-filtered means (filtered by include flag too)
    tier_start = stats_start + len(stats_rows) + 1
    tier_label_row = tier_start
    ws.cell(row=tier_label_row, column=4, value="Tier-filtered (include=TRUE)").font = SECTION_FONT
    ws.cell(row=tier_label_row, column=4).fill = SECTION_FILL
    for i, tier in enumerate([1, 2, 3], start=1):
        r = tier_start + i
        ws.cell(row=r, column=4, value=f"Tier {tier} mean").font = NORMAL_FONT
        for j, (mlabel, fmt) in enumerate(metric_columns):
            col = 5 + j
            data_range = f"{get_column_letter(col)}{first_data_row}:{get_column_letter(col)}{last_data_row}"
            tier_range = f"$A${first_data_row}:$A${last_data_row}"
            include_range = f"$B${first_data_row}:$B${last_data_row}"
            f = (
                f"=IFERROR(AVERAGEIFS({data_range},"
                f"{tier_range},{tier},"
                f"{include_range},TRUE),\"n/a\")"
            )
            ws.cell(row=r, column=col, value=f)
            ws.cell(row=r, column=col).number_format = fmt

    # Stash median/Q1/Q3 row numbers for downstream (Comps + Football Field)
    ws._stat_rows = {  # type: ignore[attr-defined]
        "max": stats_start,
        "q3": stats_start + 1,
        "median": stats_start + 2,
        "mean": stats_start + 3,
        "q1": stats_start + 4,
        "min": stats_start + 5,
    }
    ws._first_data_row = first_data_row  # type: ignore[attr-defined]
    ws._last_data_row = last_data_row  # type: ignore[attr-defined]


def build_coco_multiples_formulas(ws):
    metric_cols = [
        ("EV/Sales LTM", "0.00"),
        ("EV/Sales NTM", "0.00"),
        ("EV/EBITDA LTM", "0.0"),
        ("EV/EBITDA NTM", "0.0"),
        ("P/E LTM", "0.0"),
        ("P/E NTM", "0.0"),
    ]
    build_coco_metric_sheet(ws, "CoCo Multiples — trading multiples per comparable", metric_cols)


def build_coco_margins_formulas(ws):
    metric_cols = [
        ("Gross margin", "0.0%"),
        ("EBIT margin", "0.0%"),
        ("Net margin", "0.0%"),
    ]
    build_coco_metric_sheet(ws, "CoCo Margins — profitability per comparable", metric_cols)


def build_coco_ratios_formulas(ws):
    metric_cols = [
        ("ROE", "0.0%"),
        ("ROA", "0.0%"),
        ("D/E", "0.00"),
        ("Current ratio", "0.00"),
    ]
    build_coco_metric_sheet(ws, "CoCo Ratios — financial ratios per comparable", metric_cols)


# Deterministic row positions on CoCo Multiples — must match build_coco_metric_sheet
_COCO_FIRST_DATA_ROW = 6
_COCO_LAST_DATA_ROW = _COCO_FIRST_DATA_ROW + COCO_TABLE_ROWS - 1  # 35
_COCO_MAX_ROW = _COCO_LAST_DATA_ROW + 2  # 37
_COCO_Q3_ROW = _COCO_MAX_ROW + 1
_COCO_MEDIAN_ROW = _COCO_MAX_ROW + 2
_COCO_MEAN_ROW = _COCO_MAX_ROW + 3
_COCO_Q1_ROW = _COCO_MAX_ROW + 4
_COCO_MIN_ROW = _COCO_MAX_ROW + 5

# CoCo Multiples metric column positions (E..J)
_MULT_COL = {
    "ev_sales_ltm": "E",
    "ev_sales_ntm": "F",
    "ev_ebitda_ltm": "G",
    "ev_ebitda_ntm": "H",
    "pe_ltm": "I",
    "pe_ntm": "J",
}


def build_comps_formulas(ws):
    """Comps cross-check: apply CoCo medians (and Q1/Q3 for range) to target Y1
    metrics from Projections. Compare implied EV/Equity vs DCF result."""
    write_header_band(ws, 1, "Comps — comparable companies cross-check")
    ws.cell(row=2, column=1, value=(
        "Apply CoCo NTM trading multiples to target Y1 metrics. Implied EV from "
        "EV/Sales × Revenue and EV/EBITDA × EBITDA cross-check the DCF. P/E × Net "
        "Income gives implied Equity (bridge to EV via surplus/net debt)."
    )).font = SMALL_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

    HDR = 4
    headers = [
        "Multiple",
        "Target metric (Y1)",
        "CoCo Q1",
        "CoCo Median",
        "CoCo Q3",
        "Implied (low)",
        "Implied (mid)",
        "Implied (high)",
    ]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=HDR, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = BORDER
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    for col in ("C", "D", "E"):
        ws.column_dimensions[col].width = 14
    for col in ("F", "G", "H"):
        ws.column_dimensions[col].width = 16

    def section(row, title):
        ws.cell(row=row, column=1, value=title).font = SECTION_FONT
        ws.cell(row=row, column=1).fill = SECTION_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)

    # ----- NTM-based multiples (forward-looking; use Y1 from Projections) -----
    section(6, "NTM-based multiples (applied to Y1)")
    R_SALES, R_EBITDA, R_PE = 7, 8, 9
    multiple_rows = [
        # (label, target_cell_on_projections, multiples_col_on_CoCoMult, gives_eq)
        (R_SALES, "EV/Sales NTM", "Projections!D8", _MULT_COL["ev_sales_ntm"], False),
        (R_EBITDA, "EV/EBITDA NTM", "Projections!D15", _MULT_COL["ev_ebitda_ntm"], False),
        (R_PE, "P/E NTM (gives Equity)", "Projections!D23", _MULT_COL["pe_ntm"], True),
    ]
    for r, label, target_ref, mult_col, gives_eq in multiple_rows:
        ws.cell(row=r, column=1, value=label).font = NORMAL_FONT
        ws.cell(row=r, column=2, value=f"={target_ref}").number_format = "#,##0"
        ws.cell(row=r, column=3, value=f"='CoCo Multiples'!{mult_col}{_COCO_Q1_ROW}").number_format = (
            "0.0" if "EBITDA" in label or "P/E" in label else "0.00"
        )
        ws.cell(row=r, column=4, value=f"='CoCo Multiples'!{mult_col}{_COCO_MEDIAN_ROW}").number_format = (
            "0.0" if "EBITDA" in label or "P/E" in label else "0.00"
        )
        ws.cell(row=r, column=5, value=f"='CoCo Multiples'!{mult_col}{_COCO_Q3_ROW}").number_format = (
            "0.0" if "EBITDA" in label or "P/E" in label else "0.00"
        )
        # Implied = target × multiple
        for col_letter, mult_col_letter in (("F", "C"), ("G", "D"), ("H", "E")):
            cell = ws[f"{col_letter}{r}"]
            cell.value = f"=IFERROR(B{r}*{mult_col_letter}{r},\"n/a\")"
            cell.number_format = "#,##0"
        for col_letter in ("B", "C", "D", "E", "F", "G", "H"):
            ws[f"{col_letter}{r}"].border = BORDER

    # ----- LTM-based multiples (note — needs LTM target metrics, not yet wired) -----
    section(11, "LTM-based multiples (not wired in v1 — needs Historical FS link)")
    note_row = 12
    ws.cell(row=note_row, column=1, value=(
        "TODO: when Historical FS sheet is built, wire LTM target metrics here "
        "and apply 'CoCo Multiples'!E/G/I (LTM) columns."
    )).font = SMALL_FONT
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=8)

    # ----- Cross-check summary -----
    section(14, "Cross-check summary — DCF vs CoCo-implied")
    R_DCF_PM, R_DCF_IN, R_IMP_SALES, R_IMP_EBITDA, R_IMP_PE_EQ, R_IMP_PE_EV, R_IMP_AVG = 15, 16, 17, 18, 19, 20, 21
    cross_check = [
        (R_DCF_PM, "DCF EV — Per Management", "='DCF'!I27", "EV"),
        (R_DCF_IN, "DCF EV — Independent", "='DCF (Independent)'!I27", "EV"),
        (R_IMP_SALES, "CoCo-implied EV — EV/Sales NTM (Median)", f"=G{R_SALES}", "EV"),
        (R_IMP_EBITDA, "CoCo-implied EV — EV/EBITDA NTM (Median)", f"=G{R_EBITDA}", "EV"),
        (R_IMP_PE_EQ, "CoCo-implied Equity — P/E NTM (Median)", f"=G{R_PE}", "Equity"),
        # Convert implied Equity (P/E) to implied EV via inverse bridge:
        # EV = Equity − surplus − non_op + net_debt_override(or 0) + minority_interests
        (R_IMP_PE_EV,
         "CoCo-implied EV — P/E NTM (bridged from Equity)",
         f"=G{R_PE}-surplus_assets-non_operating_assets"
         f"+IF(ISNUMBER(net_debt_override),net_debt_override,0)+minority_interests",
         "EV"),
        (R_IMP_AVG, "Mean of CoCo-implied EVs (Sales + EBITDA + P/E)",
         f"=AVERAGE(B{R_IMP_SALES},B{R_IMP_EBITDA},B{R_IMP_PE_EV})", "EV"),
    ]
    for r, label, formula, kind in cross_check:
        ws.cell(row=r, column=1, value=label).font = NORMAL_FONT
        ws.cell(row=r, column=2, value=formula).number_format = "#,##0"
        ws.cell(row=r, column=2).border = BORDER
        if "Mean" in label or "DCF" in label:
            ws.cell(row=r, column=2).font = Font(bold=True)
            ws.cell(row=r, column=1).font = Font(bold=True)


def build_football_field_formulas(ws):
    """Football Field — methodology range comparison + weighted valuation."""
    write_header_band(ws, 1, "Football Field — methodology range reconciliation")
    ws.cell(row=2, column=1, value=(
        "Range chart input. DCF range comes from the two WACC scenarios. Comps range "
        "from CoCo Q1/Median/Q3. Precedent range from precedents_table (when populated). "
        "Apply football_field weights to compute weighted-average EV."
    )).font = SMALL_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)

    HDR = 4
    headers = ["Methodology", "Low", "Mid", "High", "Weight", "Weighted Mid", "Notes"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=HDR, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER
    ws.column_dimensions["A"].width = 36
    for col in ("B", "C", "D"):
        ws.column_dimensions[col].width = 14
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 30

    def section(row, title):
        ws.cell(row=row, column=1, value=title).font = SECTION_FONT
        ws.cell(row=row, column=1).fill = SECTION_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)

    def lrow(row, label, low, mid, high, weight=None, notes=""):
        ws.cell(row=row, column=1, value=label).font = NORMAL_FONT
        ws.cell(row=row, column=2, value=low).number_format = "#,##0"
        ws.cell(row=row, column=3, value=mid).number_format = "#,##0"
        ws.cell(row=row, column=4, value=high).number_format = "#,##0"
        if weight is not None:
            ws.cell(row=row, column=5, value=weight).number_format = "0.0%"
            ws.cell(row=row, column=6, value=f"=C{row}*E{row}").number_format = "#,##0"
        if notes:
            ws.cell(row=row, column=7, value=notes).font = SMALL_FONT
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = BORDER

    # ----- Methodology rows -----
    section(6, "Methodology ranges (EV in currency × unit)")
    R_DCF, R_C_SALES, R_C_EBITDA, R_C_PE, R_C_COMBO, R_PREC = 7, 8, 9, 10, 11, 12

    lrow(R_DCF, "DCF (across scenarios)",
         "=MIN('DCF'!I27,'DCF (Independent)'!I27)",
         "=AVERAGE('DCF'!I27,'DCF (Independent)'!I27)",
         "=MAX('DCF'!I27,'DCF (Independent)'!I27)",
         "=weight_dcf",
         "Low = Per-Mgmt vs Indep min; High = max; Mid = average")
    ws.cell(row=R_DCF, column=1).font = Font(bold=True)

    lrow(R_C_SALES, "  Comps — EV/Sales NTM",
         "=Comps!F7", "=Comps!G7", "=Comps!H7", None,
         "Reference; weighted via combined Comps row below")
    lrow(R_C_EBITDA, "  Comps — EV/EBITDA NTM",
         "=Comps!F8", "=Comps!G8", "=Comps!H8", None,
         "Reference")
    lrow(R_C_PE, "  Comps — P/E NTM (EV after bridge)",
         f"=Comps!F9-surplus_assets-non_operating_assets+IF(ISNUMBER(net_debt_override),net_debt_override,0)+minority_interests",
         f"=Comps!G9-surplus_assets-non_operating_assets+IF(ISNUMBER(net_debt_override),net_debt_override,0)+minority_interests",
         f"=Comps!H9-surplus_assets-non_operating_assets+IF(ISNUMBER(net_debt_override),net_debt_override,0)+minority_interests",
         None,
         "Reference; bridged from Equity")

    # Combined Comps row uses the AVERAGE of the three methodology mids,
    # with min/max across mids for the range.
    lrow(R_C_COMBO, "Comps — combined (mean of 3 methodologies)",
         f"=MIN(C{R_C_SALES},C{R_C_EBITDA},C{R_C_PE})",
         f"=AVERAGE(C{R_C_SALES},C{R_C_EBITDA},C{R_C_PE})",
         f"=MAX(C{R_C_SALES},C{R_C_EBITDA},C{R_C_PE})",
         "=weight_comps",
         "Low/High across the 3 mids; Mid is mean")
    ws.cell(row=R_C_COMBO, column=1).font = Font(bold=True)

    # Precedent row — uses the precedents_table if populated. If empty, show 0.
    # Mean EV/EBITDA from included precedents × Y1 EBITDA from Projections.
    # precedents_table is at Inputs!A{prec_first_data_row}:I{prec_first_data_row+14}
    # cols: include(A), date(B), acquirer(C), target(D), ev(E), ev_revenue(F), ev_ebitda(G), premium(H), rationale(I)
    # The actual rows depend on Inputs layout; use named range bounds.
    PREC_FIRST = 158  # follows cocos_table at 127:156 + 1 banner row (157) + header (157) — actual computed at runtime
    # Use a dynamic reference via the named range to avoid hardcoding
    lrow(R_PREC, "Precedent transactions",
         f"=IFERROR(MIN(IF(INDEX(precedents_table,0,1)=TRUE,INDEX(precedents_table,0,7)*Projections!D15,\"\")),0)",
         f"=IFERROR(AVERAGEIFS(INDEX(precedents_table,0,7),INDEX(precedents_table,0,1),TRUE)*Projections!D15,0)",
         f"=IFERROR(MAX(IF(INDEX(precedents_table,0,1)=TRUE,INDEX(precedents_table,0,7)*Projections!D15,\"\")),0)",
         "=weight_precedent",
         "Mean EV/EBITDA × Y1 EBITDA across included precedents")
    ws.cell(row=R_PREC, column=1).font = Font(bold=True)

    # ----- Weighted average -----
    section(14, "Weighted-average valuation")
    R_TOTAL_W, R_WAVG, R_SEL_LOW, R_SEL_MID, R_SEL_HIGH = 15, 16, 18, 19, 20
    ws.cell(row=R_TOTAL_W, column=1, value="Total weight applied").font = NORMAL_FONT
    ws.cell(row=R_TOTAL_W, column=5, value=f"=E{R_DCF}+E{R_C_COMBO}+E{R_PREC}").number_format = "0.0%"
    ws.cell(row=R_WAVG, column=1, value="Weighted-average EV (calculated)").font = Font(bold=True)
    ws.cell(row=R_WAVG, column=6,
            value=f"=IFERROR((F{R_DCF}+F{R_C_COMBO}+F{R_PREC})/E{R_TOTAL_W},0)").number_format = "#,##0"
    ws.cell(row=R_WAVG, column=6).font = Font(bold=True)

    section(17, "Selected valuation (manual override)")
    ws.cell(row=R_SEL_LOW, column=1, value="Selected low").font = NORMAL_FONT
    ws.cell(row=R_SEL_LOW, column=2, value="=selected_low").number_format = "#,##0"
    ws.cell(row=R_SEL_MID, column=1, value="Selected mid").font = NORMAL_FONT
    ws.cell(row=R_SEL_MID, column=3, value="=selected_mid").number_format = "#,##0"
    ws.cell(row=R_SEL_HIGH, column=1, value="Selected high").font = NORMAL_FONT
    ws.cell(row=R_SEL_HIGH, column=4, value="=selected_high").number_format = "#,##0"
    for r in (R_SEL_LOW, R_SEL_MID, R_SEL_HIGH):
        for c in range(1, 8):
            ws.cell(row=r, column=c).border = BORDER


def build_sensitivity_formulas(ws):
    """7x7 sensitivity grid: WACC (rows) × terminal growth rate (cols).
    Each cell recomputes Enterprise Value from scratch using its WACC and g,
    reading FCFF Y1..Y5 from Projections row 36."""
    write_header_band(ws, 1, "Sensitivity — WACC × Terminal growth rate")
    ws.cell(row=2, column=1, value=(
        "Each cell recomputes EV = Σ FCFF_y / (1+WACC)^y + FCFF_y5 × (1+g) / (WACC−g) / (1+WACC)^5. "
        "Step sizes from sens_wacc_step / sens_terminal_g_step on Inputs. Base case is "
        "current wacc_per_mgmt × terminal_growth_rate."
    )).font = SMALL_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)

    # Step counts: 7 rows × 7 cols, base in centre (index 3)
    GRID_W = 7
    BASE_IDX = 3  # 0,1,2,3,4,5,6 → centre is 3

    # Row 4: g header banner
    ws.cell(row=4, column=1, value="Sensitivity grid: rows = WACC; columns = Terminal growth rate").font = SECTION_FONT
    ws.cell(row=4, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=10)

    # Row 6: column headers — terminal growth rates
    HDR_ROW = 6
    ws.cell(row=HDR_ROW, column=2, value="WACC ↓ / g →").font = HEADER_FONT
    ws.cell(row=HDR_ROW, column=2).alignment = Alignment(horizontal="center")
    for j in range(GRID_W):
        col = 3 + j  # C..I
        offset = j - BASE_IDX  # -3..+3
        # g_value = terminal_growth_rate + offset * sens_terminal_g_step
        formula = f"=terminal_growth_rate+({offset})*sens_terminal_g_step"
        ws.cell(row=HDR_ROW, column=col, value=formula).number_format = "0.00%"
        ws.cell(row=HDR_ROW, column=col).font = HEADER_FONT
        ws.cell(row=HDR_ROW, column=col).fill = HEADER_FILL
        ws.cell(row=HDR_ROW, column=col).alignment = Alignment(horizontal="center")
        ws.cell(row=HDR_ROW, column=col).border = BORDER
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 16
    for c in range(3, 3 + GRID_W):
        ws.column_dimensions[get_column_letter(c)].width = 13

    # Rows 7-13: WACC (Per-Mgmt base ± offset × sens_wacc_step)
    BASE_FCFF_ROW = 36  # Projections row holding FCFF
    for i in range(GRID_W):
        row = HDR_ROW + 1 + i
        offset = i - BASE_IDX
        # WACC for this row
        wacc_formula = f"=wacc_per_mgmt+({offset})*sens_wacc_step"
        ws.cell(row=row, column=2, value=wacc_formula).number_format = "0.00%"
        ws.cell(row=row, column=2).font = HEADER_FONT
        ws.cell(row=row, column=2).fill = HEADER_FILL
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2).border = BORDER

        for j in range(GRID_W):
            col_letter = get_column_letter(3 + j)
            wacc_cell = f"$B${row}"
            g_cell = f"{col_letter}${HDR_ROW}"
            # EV formula: Σ FCFF_y / (1+W)^y for y=1..5 + FCFF_5 × (1+g)/(W-g) / (1+W)^5
            terms = []
            for y in range(1, PROJECTION_YEARS + 1):
                fcff_col = get_column_letter(3 + y)  # D..H
                terms.append(f"Projections!{fcff_col}{BASE_FCFF_ROW}/(1+{wacc_cell})^{y}")
            sum_explicit = "+".join(terms)
            terminal = (
                f"Projections!H{BASE_FCFF_ROW}*(1+{g_cell})/"
                f"({wacc_cell}-{g_cell})/(1+{wacc_cell})^{PROJECTION_YEARS}"
            )
            cell_formula = f"=IFERROR({sum_explicit}+{terminal},\"err\")"
            cell = ws.cell(row=row, column=3 + j, value=cell_formula)
            cell.number_format = "#,##0"
            cell.border = BORDER
            # Highlight base case (row 7+3 = row 10, col 3+3 = col F)
            if i == BASE_IDX and j == BASE_IDX:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

    # Footer note
    note_row = HDR_ROW + 1 + GRID_W + 1
    ws.cell(row=note_row, column=1, value=(
        "Highlighted cell = base case (current wacc_per_mgmt and terminal_growth_rate). "
        "EV in (currency × unit). Values do not include the EV→Equity bridge — apply "
        "surplus / debt / DLOM / DLOC manually for equity sensitivity."
    )).font = SMALL_FONT
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=10)


def build_placeholder_sheet(ws, description: str):
    write_header_band(ws, 1, ws.title)
    ws.cell(row=2, column=1, value=description).font = SMALL_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
    ws.cell(row=4, column=1, value="TODO — formulas to be added in subsequent build pass.").font = SMALL_FONT
    ws.cell(row=4, column=1).fill = TODO_FILL
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=8)
    ws.column_dimensions["A"].width = 32
    for col in range(2, 9):
        ws.column_dimensions[get_column_letter(col)].width = 16


def build_dropdowns_sheet(ws):
    """Hidden sheet holding dropdown lists for data validation (out-of-band reference)."""
    ws.title = "_dropdowns"
    ws.sheet_state = "hidden"
    lists = {
        "report_purpose": ["BEV review", "IPO pricing", "Fairness opinion", "M&A", "Fundraising"],
        "accounting_standard": ["IFRS", "US GAAP", "IFRS 9 + IFRS 13"],
        "tax_type": ["flat", "two_tier", "progressive"],
        "terminal_method": ["gordon_growth", "exit_multiple"],
        "currency": ["USD", "HKD", "SGD", "MYR", "CNY", "JPY", "KRW", "EUR", "GBP", "AUD"],
        "unit": ["'000", "million", "actual"],
        "coco_tier": ["1", "2", "3", "Excluded"],
        "revenue_growth_method": ["flat", "declining", "staged", "per_year"],
    }
    col = 1
    for name, values in lists.items():
        ws.cell(row=1, column=col, value=name).font = HEADER_FONT
        for i, v in enumerate(values, 2):
            ws.cell(row=i, column=col, value=v)
        col += 1


def build():
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    # Remove the default sheet — we add ours explicitly
    default = wb.active
    wb.remove(default)

    defined_names: list[tuple[str, str]] = []

    formula_builders = {
        "Projections": build_projections_formulas,
        "WACC": build_wacc_formulas,
        "DCF": lambda ws: build_dcf_formulas(ws, "per_mgmt"),
        "DCF (Independent)": lambda ws: build_dcf_formulas(ws, "indep"),
        "Adjustments": build_adjustments_formulas,
        "Valuation Summary": build_valuation_summary_formulas,
        "CoCo Selection": build_coco_selection_formulas,
        "CoCo Multiples": build_coco_multiples_formulas,
        "CoCo Margins": build_coco_margins_formulas,
        "CoCo Ratios": build_coco_ratios_formulas,
        "Comps": build_comps_formulas,
        "Football Field": build_football_field_formulas,
        "Sensitivity": build_sensitivity_formulas,
    }

    for name, description in SHEETS:
        ws = wb.create_sheet(title=name)
        if name == "Inputs":
            build_inputs_sheet(ws, defined_names)
        elif name in formula_builders:
            formula_builders[name](ws)
        else:
            build_placeholder_sheet(ws, description)

    dropdowns = wb.create_sheet(title="_dropdowns")
    build_dropdowns_sheet(dropdowns)

    # Apply named ranges (workbook-scope)
    for nname, ref in defined_names:
        try:
            wb.defined_names[nname] = DefinedName(name=nname, attr_text=ref)
        except (TypeError, AttributeError):
            # openpyxl >= 3.1 uses a dict-like api; older versions use append()
            wb.defined_names.append(DefinedName(name=nname, attr_text=ref))

    # Post-pass: now that named ranges exist + sheets are built, wire calculated
    # outputs from WACC back into the Inputs sheet's calculated cells.
    wire_calculated_inputs(wb)

    wb.save(OUTPUT)
    return defined_names


if __name__ == "__main__":
    names = build()
    print(f"Wrote {OUTPUT}")
    print(f"Sheets: {len(SHEETS) + 1} (incl. _dropdowns)")
    print(f"Named ranges defined: {len(names)}")
