"""Populate the Orionmano valuation workpaper skeleton from a JSON inputs file.

Reads a JSON object conforming to the inputs-sheet-schema contract, opens the
skeleton xlsx, writes every parameter into the matching named range, fills the
CoCo and Precedent tabular blocks, writes source/detail/notes into the audit
trail columns, runs validation, and saves the populated workpaper.

Run:
    python3 backend/valuation/export_workpaper.py \\
        --json inputs.json \\
        --output out/valuation.xlsx

If --skeleton is omitted, builds a fresh skeleton via build_skeleton.py first.
"""
from __future__ import annotations

import argparse
import json
import shutil
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook.defined_name import DefinedName

REPO_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_SKELETON = REPO_ROOT / "materials" / "templates" / "orionmano-valuation-template-v1.xlsx"
DEFAULT_OUTPUT = REPO_ROOT / "materials" / "templates" / "out" / "valuation.xlsx"


# ---------------------------------------------------------------------------
# JSON path mapping
# ---------------------------------------------------------------------------

# Flat scalar params: named_range_id -> json dotted path
SCALAR_PATHS: dict[str, str] = {
    # Engagement
    "company_name": "engagement.company_name",
    "company_country": "engagement.company_country",
    "company_industry_us": "engagement.company_industry_us",
    "company_industry_global": "engagement.company_industry_global",
    "valuation_date": "engagement.valuation_date",
    "report_purpose": "engagement.report_purpose",
    "accounting_standard": "engagement.accounting_standard",
    "engagement_team_partner": "engagement.engagement_team.partner",
    "engagement_team_manager": "engagement.engagement_team.manager",
    "engagement_department": "engagement.engagement_team.department",
    "client_name": "engagement.client_name",
    # Currency
    "currency": "currency.primary",
    "unit": "currency.unit",
    "currency_alt": "currency.alt",
    "fx_rate_alt": "currency.fx_rate_alt",
    # Tax
    "tax_jurisdiction": "tax.jurisdiction",
    "tax_type": "tax.type",
    "tax_rate_low": "tax.rate_low",
    "tax_rate_high": "tax.rate_high",
    "tax_threshold": "tax.threshold",
    "tax_effective_rate": "tax.effective_rate_override",
    # Projections — scalars
    "projection_years": "projections.years",
    "revenue_growth_method": "projections.revenue_growth_method",
    "revenue_y0": "projections.revenue_y0",
    "nwc_y0": "projections.nwc_y0",
    # Terminal
    "terminal_method": "terminal.method",
    "terminal_growth_rate": "terminal.growth_rate",
    "terminal_exit_multiple_type": "terminal.exit_multiple_type",
    "terminal_exit_multiple_value": "terminal.exit_multiple_value",
    # WACC shared
    "risk_free_rate": "wacc.shared.risk_free_rate",
    "risk_free_rate_source": "wacc.shared.risk_free_rate_source",
    "equity_risk_premium": "wacc.shared.equity_risk_premium",
    "country_risk_premium": "wacc.shared.country_risk_premium",
    # Bridge
    "surplus_assets": "bridge.surplus_assets",
    "net_debt_override": "bridge.net_debt_override",
    "minority_interests": "bridge.minority_interests",
    "non_operating_assets": "bridge.non_operating_assets",
    "dlom_pct": "bridge.dlom_pct",
    "dloc_pct": "bridge.dloc_pct",
    "equity_interest_pct": "bridge.equity_interest_pct",
    "shares_outstanding": "bridge.shares_outstanding",
    "shares_outstanding_diluted": "bridge.shares_outstanding_diluted",
    "pre_money_pct": "bridge.pre_money_pct",
    # Adjustments
    "capitalize_rd": "adjustments.capitalize_rd",
    "rd_amortization_years": "adjustments.rd_amortization_years",
    "convert_operating_leases": "adjustments.convert_operating_leases",
    "lease_discount_rate": "adjustments.lease_discount_rate",
    # Football field
    "weight_dcf": "football_field.weight_dcf",
    "weight_comps": "football_field.weight_comps",
    "weight_precedent": "football_field.weight_precedent",
    "weight_nav": "football_field.weight_nav",
    "selected_low": "football_field.selected_low",
    "selected_mid": "football_field.selected_mid",
    "selected_high": "football_field.selected_high",
    # Sensitivity
    "sens_wacc_step": "sensitivity.wacc_step",
    "sens_wacc_count": "sensitivity.wacc_count",
    "sens_terminal_g_step": "sensitivity.terminal_g_step",
    "sens_terminal_g_count": "sensitivity.terminal_g_count",
    "sens_revenue_g_step": "sensitivity.revenue_g_step",
    "sens_ebitda_margin_step": "sensitivity.ebitda_margin_step",
}

# Year-vector params: stem -> json dotted path (list)
YEAR_VECTOR_PATHS: dict[str, str] = {
    "revenue_growth": "projections.revenue_growth",
    "gross_margin": "projections.gross_margin",
    "opex_pct_revenue": "projections.opex_pct_revenue",
    "capex_pct_revenue": "projections.capex_pct_revenue",
    "dep_pct_revenue": "projections.dep_pct_revenue",
    "nwc_pct_sales": "projections.nwc_pct_sales",
}

# Scenario-sensitive params: stem -> (per_mgmt path, indep path)
SCENARIO_PATHS: dict[str, tuple[str, str]] = {
    "unlevered_beta": ("wacc.per_management.unlevered_beta", "wacc.independent.unlevered_beta"),
    "target_debt_to_equity": ("wacc.per_management.target_debt_to_equity", "wacc.independent.target_debt_to_equity"),
    "size_premium": ("wacc.per_management.size_premium", "wacc.independent.size_premium"),
    "specific_risk_premium": ("wacc.per_management.specific_risk_premium", "wacc.independent.specific_risk_premium"),
    "pretax_cost_of_debt": ("wacc.per_management.pretax_cost_of_debt", "wacc.independent.pretax_cost_of_debt"),
    "target_debt_weight": ("wacc.per_management.target_debt_weight", "wacc.independent.target_debt_weight"),
    "target_equity_weight": ("wacc.per_management.target_equity_weight", "wacc.independent.target_equity_weight"),
    # Calculated scenarios — left blank by exporter; downstream sheets compute
    # levered_beta, cost_of_equity, aftertax_cost_of_debt, wacc
}

# CoCo column order matches build_skeleton.py COCO_COLUMNS
COCO_FIELDS = [
    "tier", "include", "company", "ticker", "country", "accounting",
    "market_cap_usd_mm", "d_to_e", "raw_beta", "tax_rate", None,  # last is calculated
]
PRECEDENT_FIELDS = [
    "include", "date", "acquirer", "target", "ev_usd_mm",
    "ev_revenue", "ev_ebitda", "premium", "rationale",
]

# Historical FS row map — column letters: C=FY-5, D=FY-4, E=FY-3, F=FY-2, G=FY-1
HISTORICAL_FS_ROWS: dict[str, int] = {
    # Income statement
    "revenue": 7,
    "cogs": 8,
    "gross_profit": 9,
    "opex_total": 10,
    "sga": 11,
    "rnd": 12,
    "ebitda": 13,
    "da": 14,
    "ebit": 15,
    "interest_expense": 16,
    "other_income_expense": 17,
    "profit_before_tax": 18,
    "tax_expense": 19,
    "net_income": 20,
    # Balance sheet — current assets
    "cash": 24,
    "accounts_receivable": 25,
    "inventory": 26,
    "prepaid_expenses": 27,
    "total_current_assets": 28,
    # Non-current assets
    "ppe": 30,
    "intangibles": 31,
    "other_lt_assets": 32,
    "total_assets": 33,
    # Current liabilities
    "accounts_payable": 36,
    "short_term_debt": 37,
    "other_current_liabilities": 38,
    "total_current_liabilities": 39,
    # Non-current liabilities
    "long_term_debt": 41,
    "other_lt_liabilities": 42,
    "total_liabilities": 43,
    "total_equity": 44,
}

# CoCo metric sheet layout — first_data_row=6, last_data_row=35 (30 rows)
# Cols start at 5 (E). Order matches build_coco_metric_sheet metric_columns.
COCO_METRIC_LAYOUT: dict[str, dict[str, Any]] = {
    "CoCo Multiples": {
        "json_key": "coco_multiples",
        "fields": ["ev_sales_ltm", "ev_sales_ntm", "ev_ebitda_ltm",
                   "ev_ebitda_ntm", "pe_ltm", "pe_ntm"],
    },
    "CoCo Margins": {
        "json_key": "coco_margins",
        "fields": ["gross", "ebit", "net"],
    },
    "CoCo Ratios": {
        "json_key": "coco_ratios",
        "fields": ["roe", "roa", "d_to_e", "current_ratio"],
    },
}
COCO_METRIC_FIRST_ROW = 6
COCO_METRIC_FIRST_COL = 5
COCO_METRIC_MAX_ROWS = 30


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------

@dataclass
class ValidationResult:
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    def err(self, msg: str) -> None:
        self.errors.append(msg)

    def warn(self, msg: str) -> None:
        self.warnings.append(msg)

    @property
    def ok(self) -> bool:
        return not self.errors


def validate_payload(payload: dict, vr: ValidationResult) -> None:
    # Required engagement fields
    for path in ("engagement.company_name", "engagement.valuation_date",
                 "currency.primary", "currency.unit"):
        if get_path(payload, path) in (None, ""):
            vr.err(f"required field missing: {path}")

    # Football field weights must sum to 1.0 if any are set
    weights = [get_path(payload, f"football_field.weight_{k}", 0)
               for k in ("dcf", "comps", "precedent", "nav")]
    if any(w not in (None, 0) for w in weights):
        total = sum(w or 0 for w in weights)
        if abs(total - 1.0) > 0.001:
            vr.err(f"football_field weights must sum to 1.0, got {total:.4f}")

    # Percentage bounds
    for path, lo, hi, name in [
        ("bridge.dlom_pct", 0, 0.5, "DLOM"),
        ("bridge.dloc_pct", 0, 0.5, "DLOC"),
        ("terminal.growth_rate", 0, 0.05, "terminal growth"),
        ("wacc.shared.risk_free_rate", 0, 0.15, "risk-free rate"),
        ("wacc.shared.equity_risk_premium", 0.04, 0.10, "ERP"),
    ]:
        v = get_path(payload, path)
        if v is None:
            continue
        if v < lo or v > hi:
            vr.warn(f"{name} ({path}={v}) outside typical range [{lo}, {hi}]")

    # Tax rates plausibility
    for path in ("tax.rate_low", "tax.rate_high", "tax.effective_rate_override"):
        v = get_path(payload, path)
        if v is None:
            continue
        if v < 0 or v > 0.5:
            vr.err(f"{path}={v} not a plausible tax rate (0..0.5)")

    # CoCo / Precedent capacity
    cocos = payload.get("cocos") or []
    if len(cocos) > 30:
        vr.warn(f"cocos has {len(cocos)} rows; only first 30 will be written")
    precedents = payload.get("precedents") or []
    if len(precedents) > 15:
        vr.warn(f"precedents has {len(precedents)} rows; only first 15 will be written")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def get_path(obj: dict, dotted: str, default: Any = None) -> Any:
    """Walk a dotted JSON path. Returns default if any step missing."""
    cur: Any = obj
    for part in dotted.split("."):
        if isinstance(cur, dict) and part in cur:
            cur = cur[part]
        else:
            return default
    return cur


def parse_named_range(dn: DefinedName) -> tuple[str, str] | None:
    """Resolve a workbook-level DefinedName to (sheet_name, cell_address). Returns None for ranges."""
    # attr_text e.g. "'Inputs'!$D$7"
    text = dn.attr_text or ""
    if ":" in text:
        return None  # range, not a single cell
    if "!" not in text:
        return None
    sheet_part, cell_part = text.rsplit("!", 1)
    sheet = sheet_part.strip("'")
    cell = cell_part.replace("$", "")
    return sheet, cell


def cell_addr_components(addr: str) -> tuple[str, int]:
    """Split 'D7' into ('D', 7)."""
    col = "".join(ch for ch in addr if ch.isalpha())
    row = int("".join(ch for ch in addr if ch.isdigit()))
    return col, row


# ---------------------------------------------------------------------------
# Writer
# ---------------------------------------------------------------------------

def write_value_with_audit(ws, addr: str, value: Any, source_meta: dict | None) -> None:
    """Write value at addr, plus source/detail/notes in F/G/H of same row."""
    col_letter, row = cell_addr_components(addr)
    if value is not None:
        ws[addr] = value
    if source_meta:
        # F = Source, G = Source detail, H = Notes
        if "source" in source_meta:
            ws.cell(row=row, column=6, value=source_meta["source"])
        if "detail" in source_meta:
            ws.cell(row=row, column=7, value=source_meta["detail"])
        if "notes" in source_meta:
            ws.cell(row=row, column=8, value=source_meta["notes"])


def populate_scalars(wb, payload: dict, vr: ValidationResult) -> int:
    sources = payload.get("sources", {})
    inputs_ws = wb["Inputs"]
    written = 0

    # Plain scalars
    for pid, json_path in SCALAR_PATHS.items():
        if pid not in wb.defined_names:
            vr.warn(f"named range '{pid}' not present in workbook — skipped")
            continue
        resolved = parse_named_range(wb.defined_names[pid])
        if resolved is None:
            continue
        sheet, addr = resolved
        ws = wb[sheet]
        value = get_path(payload, json_path)
        meta = sources.get(pid)
        write_value_with_audit(ws, addr, value, meta)
        written += 1

    # Year vectors
    for stem, json_path in YEAR_VECTOR_PATHS.items():
        arr = get_path(payload, json_path) or []
        for i, v in enumerate(arr, start=1):
            pid = f"{stem}_y{i}"
            if pid not in wb.defined_names:
                continue
            resolved = parse_named_range(wb.defined_names[pid])
            if resolved is None:
                continue
            sheet, addr = resolved
            ws = wb[sheet]
            meta = sources.get(pid)
            write_value_with_audit(ws, addr, v, meta)
            written += 1

    # Scenario columns
    for stem, (pm_path, indep_path) in SCENARIO_PATHS.items():
        for suffix, jpath in (("per_mgmt", pm_path), ("indep", indep_path)):
            pid = f"{stem}_{suffix}"
            if pid not in wb.defined_names:
                continue
            resolved = parse_named_range(wb.defined_names[pid])
            if resolved is None:
                continue
            sheet, addr = resolved
            ws = wb[sheet]
            value = get_path(payload, jpath)
            meta = sources.get(pid)
            write_value_with_audit(ws, addr, value, meta)
            written += 1

    return written


def populate_historical_fs(wb, payload: dict, vr: ValidationResult) -> int:
    """Write 5-year historical FS arrays into the Historical FS sheet.

    Each entry under payload['historical_fs'] is a 5-element array ordered
    FY-5..FY-1 (oldest to most recent). Writes into columns C..G (3..7).
    """
    hf = payload.get("historical_fs") or {}
    if not hf:
        return 0
    if "Historical FS" not in wb.sheetnames:
        vr.warn("Historical FS sheet missing — skipped historical_fs population")
        return 0
    ws = wb["Historical FS"]
    written = 0
    for field_name, arr in hf.items():
        if field_name not in HISTORICAL_FS_ROWS:
            vr.warn(f"historical_fs.{field_name} not recognized — skipped")
            continue
        if not isinstance(arr, list):
            vr.warn(f"historical_fs.{field_name} must be a list of 5 numbers")
            continue
        row = HISTORICAL_FS_ROWS[field_name]
        for i, v in enumerate(arr[:5]):
            if v is None:
                continue
            ws.cell(row=row, column=3 + i, value=v)
            written += 1
    return written


def populate_coco_metrics(wb, payload: dict, vr: ValidationResult) -> int:
    """Fill the CoCo Multiples / Margins / Ratios sheets from the producer's
    coco_multiples / coco_margins / coco_ratios arrays. Each array is aligned
    row-by-row with payload['cocos'] — index N here matches comparable N.
    """
    written = 0
    for sheet_name, layout in COCO_METRIC_LAYOUT.items():
        arr = payload.get(layout["json_key"]) or []
        if not arr:
            continue
        if sheet_name not in wb.sheetnames:
            vr.warn(f"{sheet_name} sheet missing — skipped {layout['json_key']}")
            continue
        ws = wb[sheet_name]
        for i, entry in enumerate(arr[:COCO_METRIC_MAX_ROWS]):
            if not isinstance(entry, dict):
                continue
            row = COCO_METRIC_FIRST_ROW + i
            for j, field_name in enumerate(layout["fields"]):
                v = entry.get(field_name)
                if v is None:
                    continue
                ws.cell(row=row, column=COCO_METRIC_FIRST_COL + j, value=v)
                written += 1
    return written


def validate_sources_completeness(payload: dict, vr: ValidationResult) -> None:
    """Warn when a scalar parameter has a value but no sources entry.

    Audit-trail discipline (Damodaran Aramco pattern): every parameter that
    drives valuation should cite a source.
    """
    sources = payload.get("sources") or {}
    high_priority = {
        "company_name", "valuation_date", "currency", "tax_rate_high",
        "revenue_y0", "nwc_y0", "risk_free_rate", "equity_risk_premium",
        "country_risk_premium", "unlevered_beta_per_mgmt", "dlom_pct",
        "dloc_pct", "shares_outstanding", "terminal_growth_rate",
    }
    missing: list[str] = []
    for pid in SCALAR_PATHS:
        if pid not in high_priority:
            continue
        json_path = SCALAR_PATHS[pid]
        v = get_path(payload, json_path)
        if v in (None, ""):
            continue
        meta = sources.get(pid)
        if not meta or not meta.get("source"):
            missing.append(pid)
    # Scenarioed high-priority too
    for stem in ("unlevered_beta", "specific_risk_premium"):
        for suffix in ("per_mgmt", "indep"):
            pid = f"{stem}_{suffix}"
            if pid in sources and sources[pid].get("source"):
                continue
            tup = SCENARIO_PATHS.get(stem)
            if not tup:
                continue
            jpath = tup[0] if suffix == "per_mgmt" else tup[1]
            if get_path(payload, jpath) in (None, ""):
                continue
            missing.append(pid)
    if missing:
        sample = ", ".join(missing[:8])
        more = f" (and {len(missing) - 8} more)" if len(missing) > 8 else ""
        vr.warn(
            f"sources missing for {len(missing)} high-priority parameters: "
            f"{sample}{more}"
        )


def populate_tables(wb, payload: dict, vr: ValidationResult) -> int:
    written = 0
    inputs_ws = wb["Inputs"]

    # CoCo table
    if "cocos_table" in wb.defined_names:
        dn = wb.defined_names["cocos_table"]
        text = dn.attr_text  # 'Inputs'!$A$N:$K$M
        sheet_part, range_part = text.rsplit("!", 1)
        start_addr, _ = range_part.split(":")
        start_addr = start_addr.replace("$", "")
        _, start_row = cell_addr_components(start_addr)
        cocos = (payload.get("cocos") or [])[:30]
        for i, coco in enumerate(cocos):
            r = start_row + i
            for col_idx, field_name in enumerate(COCO_FIELDS, start=1):
                if field_name is None:
                    continue
                v = coco.get(field_name)
                inputs_ws.cell(row=r, column=col_idx, value=v)
            written += 1

    # Precedent table
    if "precedents_table" in wb.defined_names:
        dn = wb.defined_names["precedents_table"]
        text = dn.attr_text
        sheet_part, range_part = text.rsplit("!", 1)
        start_addr, _ = range_part.split(":")
        start_addr = start_addr.replace("$", "")
        _, start_row = cell_addr_components(start_addr)
        precedents = (payload.get("precedents") or [])[:15]
        for i, p in enumerate(precedents):
            r = start_row + i
            for col_idx, field_name in enumerate(PRECEDENT_FIELDS, start=1):
                v = p.get(field_name)
                inputs_ws.cell(row=r, column=col_idx, value=v)
            written += 1

    return written


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def export(json_path: Path, skeleton_path: Path, output_path: Path) -> ValidationResult:
    if not skeleton_path.exists():
        # Auto-build skeleton if missing
        from build_skeleton import build  # type: ignore
        build()
    if not json_path.exists():
        raise FileNotFoundError(f"JSON inputs not found: {json_path}")

    with json_path.open() as f:
        payload = json.load(f)

    vr = ValidationResult()
    validate_payload(payload, vr)
    validate_sources_completeness(payload, vr)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(skeleton_path, output_path)

    wb = load_workbook(output_path)
    n_scalar = populate_scalars(wb, payload, vr)
    n_table = populate_tables(wb, payload, vr)
    n_hist = populate_historical_fs(wb, payload, vr)
    n_coco = populate_coco_metrics(wb, payload, vr)
    wb.save(output_path)

    print(f"Scalars written:    {n_scalar}")
    print(f"Table rows written: {n_table}")
    print(f"Historical FS cells: {n_hist}")
    print(f"CoCo metric cells:   {n_coco}")
    print(f"Errors:   {len(vr.errors)}")
    print(f"Warnings: {len(vr.warnings)}")
    for e in vr.errors:
        print(f"  ERROR:   {e}")
    for w in vr.warnings:
        print(f"  WARNING: {w}")
    print(f"Output: {output_path}")
    return vr


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--json", type=Path, required=True, help="Path to JSON inputs file")
    parser.add_argument("--skeleton", type=Path, default=DEFAULT_SKELETON,
                        help=f"Path to skeleton xlsx (default: {DEFAULT_SKELETON})")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT,
                        help=f"Path to write populated xlsx (default: {DEFAULT_OUTPUT})")
    args = parser.parse_args(argv)
    vr = export(args.json, args.skeleton, args.output)
    return 0 if vr.ok else 2


if __name__ == "__main__":
    sys.exit(main())
