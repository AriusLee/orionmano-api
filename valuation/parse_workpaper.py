"""Reverse of export_workpaper: read a populated/edited xlsx and reconstruct
the inputs JSON. Eric 2026-05-08 item 8 round-trip.

Strategy:
1. The export step stashes the full inputs JSON in a hidden '_meta' sheet.
   Use that as a high-fidelity BASELINE so fields not rendered in the visible
   Inputs sheet (rationales, sources, segments, schema additions) survive
   the round-trip.
2. Walk the named ranges defined in build_skeleton.SECTIONS / COCO_COLUMNS /
   PRECEDENT_COLUMNS and OVERLAY each visible-cell value onto the baseline.
   Any cell the user edited in Excel takes precedence.
3. Return the overlaid JSON. The caller validates against the Pydantic schema
   before running the rest of the pipeline.
"""

from __future__ import annotations

import json
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from openpyxl.workbook.defined_name import DefinedName

from build_skeleton import SECTIONS, COCO_COLUMNS, PRECEDENT_COLUMNS  # type: ignore


# ─── PID → JSON path map (built once from build_skeleton.SECTIONS) ────────────

# Hand-crafted mapping for fields whose PID doesn't trivially map to a JSON key
# (e.g. tax fields collapse the `tax_` prefix, engagement_team_* nests under
# engagement.engagement_team, sensitivity drops the `sens_` prefix).
_SCALAR_MAP_OVERRIDES: dict[str, list[str]] = {
    # Section A — engagement
    "engagement_team_partner": ["engagement", "engagement_team", "partner"],
    "engagement_team_manager": ["engagement", "engagement_team", "manager"],
    "engagement_department": ["engagement", "engagement_team", "department"],
    # Section B — currency
    "currency": ["currency", "primary"],
    "unit": ["currency", "unit"],
    "currency_alt": ["currency", "alt"],
    "fx_rate_alt": ["currency", "fx_rate_alt"],
    # Section C — tax (`tax_` prefix dropped in JSON)
    "tax_jurisdiction": ["tax", "jurisdiction"],
    "tax_type": ["tax", "type"],
    "tax_rate_low": ["tax", "rate_low"],
    "tax_rate_high": ["tax", "rate_high"],
    "tax_threshold": ["tax", "threshold"],
    "tax_effective_rate": ["tax", "effective_rate_override"],
    # Section D — scalars (excluding Y-vector params handled below)
    "projection_years": ["projections", "years"],
    "revenue_growth_method": ["projections", "revenue_growth_method"],
    "revenue_y0": ["projections", "revenue_y0"],
    "nwc_y0": ["projections", "nwc_y0"],
    # Section E — terminal (`terminal_` prefix dropped)
    "terminal_method": ["terminal", "method"],
    "terminal_growth_rate": ["terminal", "growth_rate"],
    "terminal_exit_multiple_type": ["terminal", "exit_multiple_type"],
    "terminal_exit_multiple_value": ["terminal", "exit_multiple_value"],
    "terminal_nominal_gdp_growth": ["terminal", "nominal_gdp_growth"],
    # Section F — WACC shared
    "risk_free_rate": ["wacc", "shared", "risk_free_rate"],
    "risk_free_rate_source": ["wacc", "shared", "risk_free_rate_source"],
    "equity_risk_premium": ["wacc", "shared", "equity_risk_premium"],
    "country_risk_premium": ["wacc", "shared", "country_risk_premium"],
    # Section L — sensitivity (`sens_` prefix dropped)
    "sens_wacc_step": ["sensitivity", "wacc_step"],
    "sens_wacc_count": ["sensitivity", "wacc_count"],
    "sens_terminal_g_step": ["sensitivity", "terminal_g_step"],
    "sens_terminal_g_count": ["sensitivity", "terminal_g_count"],
    "sens_revenue_g_step": ["sensitivity", "revenue_g_step"],
    "sens_ebitda_margin_step": ["sensitivity", "ebitda_margin_step"],
}

# Scenario fields appear in the visible sheet (per-mgmt + independent columns)
# but only a subset are JSON inputs; the rest are calculated downstream.
_SCENARIO_INPUT_FIELDS = {
    "unlevered_beta",
    "target_debt_to_equity",
    "size_premium",
    "specific_risk_premium",
    "pretax_cost_of_debt",
    "target_debt_weight",
    "target_equity_weight",
}

# Default JSON root for each Section.code (used when no override is present).
_SECTION_ROOTS = {
    "A": "engagement",
    "B": "currency",
    "C": "tax",
    "D": "projections",
    "E": "terminal",
    "F": "wacc",
    "I": "bridge",
    "J": "adjustments",
    "K": "football_field",
    "L": "sensitivity",
}

# Y-vector params: revenue_growth_y1..y5 → projections.revenue_growth[0..4].
# Detected at parse time via the `_y\d+$` suffix.
_Y_VECTOR_ROOTS = {
    "revenue_growth": ["projections", "revenue_growth"],
    "revenue_growth_primary": ["projections", "revenue_growth_primary"],
    "gross_margin": ["projections", "gross_margin"],
    "opex_pct_revenue": ["projections", "opex_pct_revenue"],
    "capex_pct_revenue": ["projections", "capex_pct_revenue"],
    "dep_pct_revenue": ["projections", "dep_pct_revenue"],
    "nwc_pct_sales": ["projections", "nwc_pct_sales"],
}


def _build_scalar_map() -> dict[str, tuple[list[str], str]]:
    """Build PID → (json_path, param_type) for every scalar Inputs cell.
    Param.type_ is used downstream to coerce the cell value correctly."""
    out: dict[str, tuple[list[str], str]] = {}
    for sect in SECTIONS:
        root = _SECTION_ROOTS.get(sect.code)
        for p in sect.params:
            # Scenario field → two named ranges (PID_per_mgmt / PID_indep)
            if p.scenario:
                if p.pid not in _SCENARIO_INPUT_FIELDS:
                    continue  # calculated field; skip
                out[f"{p.pid}_per_mgmt"] = (["wacc", "per_management", p.pid], p.type_)
                out[f"{p.pid}_indep"] = (["wacc", "independent", p.pid], p.type_)
                continue
            # Y-vector field → handled separately by _y\d+$ detection
            ym = re.match(r"^(.+)_y(\d+)$", p.pid)
            if ym and ym.group(1) in _Y_VECTOR_ROOTS:
                # Path is roots[base] + index — but we store the year index in
                # the path encoded as a special sentinel to be resolved at write.
                base = ym.group(1)
                year = int(ym.group(2))
                out[p.pid] = ([*_Y_VECTOR_ROOTS[base], f"__idx__{year - 1}"], p.type_)
                continue
            # Specific override?
            if p.pid in _SCALAR_MAP_OVERRIDES:
                out[p.pid] = (_SCALAR_MAP_OVERRIDES[p.pid], p.type_)
                continue
            # Default: under section root, key == PID
            if root:
                out[p.pid] = ([root, p.pid], p.type_)
    return out


_SCALAR_MAP = _build_scalar_map()


# ─── Type coercion ────────────────────────────────────────────────────────────

def _coerce(value: Any, type_: str) -> Any:
    """Map an Excel cell value to the right Python/JSON type. Returns None for
    blanks. Doesn't raise on bad data — callers prefer the baseline value when
    coercion fails."""
    if value is None:
        return None
    if isinstance(value, str) and not value.strip():
        return None
    if type_ in ("number", "percentage", "currency"):
        try:
            return float(value)
        except (TypeError, ValueError):
            return None
    if type_ == "boolean":
        if isinstance(value, bool):
            return value
        s = str(value).strip().lower()
        if s in ("true", "yes", "1", "y"):
            return True
        if s in ("false", "no", "0", "n"):
            return False
        return None
    if type_ == "date":
        if isinstance(value, (date, datetime)):
            return value.isoformat()[:10]
        return str(value)
    # text / enum
    return str(value)


# ─── Defined-name lookup ──────────────────────────────────────────────────────

def _cell_for_name(wb, name: str) -> Any:
    """Resolve a defined name to its single cell value, or None if the name
    isn't defined / the destination is malformed."""
    try:
        dn: DefinedName | None = wb.defined_names.get(name)
    except Exception:
        dn = None
    if dn is None:
        return None
    try:
        dests = list(dn.destinations)
    except Exception:
        return None
    if not dests:
        return None
    sheet_title, coord = dests[0]
    try:
        ws = wb[sheet_title]
    except KeyError:
        return None
    return ws[coord].value


def _range_rows(wb, name: str) -> list[list[Any]]:
    """Resolve a defined name pointing at a rectangular range; return its rows
    as 2-D list of cell values. Empty list if the name doesn't exist."""
    try:
        dn: DefinedName | None = wb.defined_names.get(name)
    except Exception:
        return []
    if dn is None:
        return []
    try:
        dests = list(dn.destinations)
    except Exception:
        return []
    if not dests:
        return []
    sheet_title, coord = dests[0]
    try:
        ws = wb[sheet_title]
    except KeyError:
        return []
    try:
        min_col, min_row, max_col, max_row = range_boundaries(coord)
    except Exception:
        return []
    rows: list[list[Any]] = []
    for r in range(min_row, max_row + 1):
        row_vals: list[Any] = []
        for c in range(min_col, max_col + 1):
            row_vals.append(ws.cell(row=r, column=c).value)
        rows.append(row_vals)
    return rows


def _set_path(obj: dict, path: list[str], value: Any) -> None:
    """Write `value` into `obj` at the given path, creating intermediate dicts
    as needed. Path entries of the form `__idx__N` mean: at this point the
    container should be a list and we write at index N."""
    cur: Any = obj
    for i, key in enumerate(path):
        last = i == len(path) - 1
        if key.startswith("__idx__"):
            idx = int(key[len("__idx__"):])
            # Container must be a list
            if not isinstance(cur, list):
                # Replace the parent's slot with a list; we walked to this point
                # via the previous key. But we don't have a back-reference, so
                # build a list of the right size at the parent.
                # Easier: require callers structure correctly — fall back to overwrite.
                return
            while len(cur) <= idx:
                cur.append(None)
            if last:
                cur[idx] = value
                return
            if not isinstance(cur[idx], (dict, list)):
                cur[idx] = {}
            cur = cur[idx]
            continue
        # Normal dict key
        if last:
            if isinstance(cur, dict):
                cur[key] = value
            return
        nxt_is_list = path[i + 1].startswith("__idx__")
        if isinstance(cur, dict):
            if key not in cur or cur[key] is None or (
                nxt_is_list and not isinstance(cur[key], list)
            ) or (
                not nxt_is_list and not isinstance(cur[key], (dict, list))
            ):
                cur[key] = [] if nxt_is_list else {}
            cur = cur[key]


# ─── Parser entry point ───────────────────────────────────────────────────────

def parse(xlsx_path: str | Path) -> dict[str, Any]:
    """Parse the given xlsx and return the inputs JSON. Uses the hidden
    `_meta` sheet (written by export_workpaper) as a baseline, then overlays
    visible-cell edits read via named ranges."""
    wb = load_workbook(xlsx_path, data_only=True)

    # 1. Baseline: hidden _meta sheet, if present.
    baseline: dict[str, Any] = {}
    if "_meta" in wb.sheetnames:
        try:
            meta_ws = wb["_meta"]
            raw = meta_ws["A2"].value
            if raw:
                baseline = json.loads(raw)
        except (json.JSONDecodeError, KeyError):
            baseline = {}

    # 2. Scalars: walk every PID in the scalar map and overlay.
    for pid, (path, type_) in _SCALAR_MAP.items():
        cell_value = _cell_for_name(wb, pid)
        coerced = _coerce(cell_value, type_)
        if coerced is None:
            continue  # blank cell → keep baseline value
        _set_path(baseline, path, coerced)

    # 3. Cocos table — overlay rows. The named range covers all 30 rows; we
    # only keep rows where the Company name (col index 2) is non-empty.
    coco_rows = _range_rows(wb, "cocos_table")
    if coco_rows:
        company_col_idx = next(
            (i for i, (cid, _, _) in enumerate(COCO_COLUMNS) if cid == "coco_company"),
            2,
        )
        kept: list[dict[str, Any]] = []
        for row in coco_rows:
            company = row[company_col_idx] if company_col_idx < len(row) else None
            if company is None or (isinstance(company, str) and not company.strip()):
                continue
            entry: dict[str, Any] = {}
            for i, (cid, _, ctype) in enumerate(COCO_COLUMNS):
                if i >= len(row):
                    break
                # JSON key drops the coco_ prefix (matches sample_inputs.json)
                json_key = cid[len("coco_"):] if cid.startswith("coco_") else cid
                v = _coerce(row[i], ctype)
                if v is not None:
                    entry[json_key] = v
            if entry:
                kept.append(entry)
        if kept:
            # Preserve baseline-only fields per comp (exchange, business_description,
            # selected_for_wacc — these aren't yet in the visible Inputs sheet)
            # by merging on company name where possible.
            baseline_cocos = baseline.get("cocos") if isinstance(baseline.get("cocos"), list) else []
            by_company = {
                (c.get("company") or "").strip().lower(): c
                for c in baseline_cocos
                if isinstance(c, dict)
            }
            merged: list[dict[str, Any]] = []
            for entry in kept:
                key = (entry.get("company") or "").strip().lower()
                base = by_company.get(key, {})
                merged.append({**base, **entry})
            baseline["cocos"] = merged

    # 4. Precedents table — overlay rows.
    prec_rows = _range_rows(wb, "precedents_table")
    if prec_rows:
        target_col_idx = next(
            (i for i, (cid, _, _) in enumerate(PRECEDENT_COLUMNS) if cid == "precedent_target"),
            3,
        )
        kept_p: list[dict[str, Any]] = []
        for row in prec_rows:
            tgt = row[target_col_idx] if target_col_idx < len(row) else None
            if tgt is None or (isinstance(tgt, str) and not tgt.strip()):
                continue
            entry = {}
            for i, (cid, _, ctype) in enumerate(PRECEDENT_COLUMNS):
                if i >= len(row):
                    break
                json_key = cid[len("precedent_"):] if cid.startswith("precedent_") else cid
                # Some PRECEDENT_COLUMNS json keys diverge from the cid stem:
                if json_key == "ev":
                    json_key = "ev_usd_mm"
                v = _coerce(row[i], ctype)
                if v is not None:
                    entry[json_key] = v
            if entry:
                kept_p.append(entry)
        if kept_p:
            baseline["precedents"] = kept_p

    # 5. Segments table (Eric 2026-05-08 item 2) — overlay rows from the
    # optional segmented revenue model. Layout: col 0=name, 1=start_year,
    # 2=initial revenue, 3-7=revenue_growth Y1-Y5, 8-12=gross_margin Y1-Y5,
    # 13-17=opex_pct_revenue Y1-Y5, 18=source, 19=growth_basis,
    # 20=contractual_support.
    seg_rows = _range_rows(wb, "segments_table")
    if seg_rows:
        kept_seg: list[dict[str, Any]] = []
        for row in seg_rows:
            name = row[0] if len(row) > 0 else None
            if name is None or (isinstance(name, str) and not name.strip()):
                continue
            entry: dict[str, Any] = {"name": str(name).strip()}
            start_year = _coerce(row[1] if len(row) > 1 else None, "number")
            if start_year is not None:
                entry["start_year"] = int(start_year)
            initial = _coerce(row[2] if len(row) > 2 else None, "number")
            if initial is not None:
                if (entry.get("start_year") or 0) == 0:
                    entry["revenue_y0"] = initial
                else:
                    entry["initial_revenue"] = initial

            def _vec(first_col: int) -> list[Any]:
                arr = [_coerce(row[first_col + k] if first_col + k < len(row) else None,
                               "percentage") for k in range(5)]
                # Trim trailing Nones — keep only the contiguous filled prefix.
                while arr and arr[-1] is None:
                    arr.pop()
                return arr

            growth = _vec(3)
            if growth:
                entry["revenue_growth"] = growth
            gm = _vec(8)
            if gm:
                entry["gross_margin"] = gm
            opex = _vec(13)
            if opex:
                entry["opex_pct_revenue"] = opex
            for col_idx, key in ((18, "source"), (19, "growth_basis"),
                                 (20, "contractual_support")):
                v = _coerce(row[col_idx] if col_idx < len(row) else None, "text")
                if v is not None:
                    entry[key] = v
            kept_seg.append(entry)
        # Only overlay when the analyst actually filled in segments — otherwise
        # we'd clobber any segments the LLM produced (which are preserved in the
        # _meta baseline at parse start). Merge on segment name so metadata the
        # visible sheet doesn't carry (or that the analyst blanked) survives from
        # the baseline.
        if kept_seg:
            baseline_segs = ((baseline.get("projections") or {}).get("segments")
                             if isinstance(baseline.get("projections"), dict) else None) or []
            by_name = {
                (s.get("name") or "").strip().lower(): s
                for s in baseline_segs if isinstance(s, dict)
            }

            def _export_view(arr: Any) -> list[Any]:
                """What the exporter wrote to the visible 5-column table for a
                baseline array: first 5 entries, trailing Nones trimmed."""
                out = list(arr[:5]) if isinstance(arr, list) else []
                while out and out[-1] is None:
                    out.pop()
                return out

            def _arrays_equal(a: list[Any], b: list[Any]) -> bool:
                if len(a) != len(b):
                    return False
                for x, y in zip(a, b):
                    if x is None or y is None:
                        if x is not y:
                            return False
                    elif abs(float(x) - float(y)) > 1e-9:
                        return False
                return True

            merged_segs = []
            for entry in kept_seg:
                base = by_name.get((entry.get("name") or "").strip().lower(), {})
                # If a visible array is byte-for-byte what the exporter wrote,
                # the analyst didn't edit it — keep the baseline's full-fidelity
                # version (the visible table caps arrays at 5 columns, which
                # truncates the 6-entry Y0-indexed arrays of start_year=0
                # segments).
                for key in ("revenue_growth", "gross_margin", "opex_pct_revenue"):
                    if key in entry and isinstance(base.get(key), list):
                        if _arrays_equal(entry[key], _export_view(base[key])):
                            entry.pop(key)
                merged_segs.append({**base, **entry})
            baseline.setdefault("projections", {})["segments"] = merged_segs

    return baseline


def main(argv: list[str] | None = None) -> int:
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", type=Path, required=True)
    ap.add_argument("--out", type=Path, default=None)
    args = ap.parse_args(argv)
    payload = parse(args.xlsx)
    text = json.dumps(payload, indent=2, default=str)
    if args.out:
        args.out.write_text(text)
        print(f"Wrote {args.out}")
    else:
        print(text)
    return 0


if __name__ == "__main__":
    import sys
    sys.exit(main())
